import copy
import difflib
import hashlib
import html
import json
import re
import secrets
import sqlite3
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from html.parser import HTMLParser
from pathlib import Path
from threading import Lock
from typing import Any
from urllib.parse import parse_qs, quote_plus, urlencode, unquote, urlparse, urlunparse

import openpyxl
import pandas as pd
import requests
import streamlit as st


APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
EXPORT_DIR = APP_DIR / "outputs"
DB_PATH = DATA_DIR / "users.sqlite3"
BUILTIN_REFERENCE_PATH = APP_DIR / "reference_data.json"
SEARCH_CACHE_PATH = DATA_DIR / "search_cache.json"

SITE_NAME = "Label Info Finder"
NEED_REVIEW_FILL = openpyxl.styles.PatternFill("solid", fgColor="FFF2CC")
NEED_REVIEW_FILL_RGBS = {"FFF2CC", "00FFF2CC", "FFFFF2CC"}
MAX_SOURCE_URLS = 3
DEFAULT_SEARCH_MODE = "Fast"
SEARCH_MODES = ["Fast", "Deep"]
CACHE_LOCK = Lock()
SEARCH_CACHE_VERSION = 2

DISTRIBUTOR = (
    "DISTRIBUTED BY / DISTRIBUÉ PAR: Nakama Trading Ltd, Scarborough, "
    "Ontario, M1X 2E5 service1@nakamatrading.com"
)
DEFAULT_DIRECTION_EN = (
    "DIRECTION FOR USE: Apply a proper amount onto lips and cheeks using a "
    "brush or fingertips. Blend evenly."
)
DEFAULT_DIRECTION_FR = (
    "MODE D’EMPLOI: Appliquer une quantité appropriée sur les lèvres et les "
    "joues à l’aide d’un pinceau ou du bout des doigts. Estomper uniformément."
)
DEFAULT_CAUTION_EN = (
    "CAUTIONS: For external use only. Discontinue use if irritation occurs. "
    "Keep out of reach of children."
)
DEFAULT_CAUTION_FR = (
    "MISES EN GARDE: Pour usage externe seulement. Cesser l'utilisation si une "
    "irritation se manifeste. Garder hors de la portée des enfants."
)
HOTLIST_URL = (
    "https://www.canada.ca/en/health-canada/services/consumer-product-safety/"
    "cosmetics/cosmetic-ingredient-hotlist-prohibited-restricted-ingredients/"
    "hotlist.html"
)
INGREDIENTS_OUTPUT_PREFIX = "INGREDIENTS/INGRÉDIENTS:"

REQUIRED_LABEL_FIELDS = [
    "product name french",
    "net weight",
    "direction for use",
    "mode d’emploi",
    "cautions",
    "mises en garde:",
    "ingredients/ingrédients",
    "distributed by / distribué par:",
    "coo",
]

SHARED_FAMILY_FIELDS = [
    "net weight",
    "direction for use",
    "mode d’emploi",
    "cautions",
    "mises en garde:",
    "distributed by / distribué par:",
    "coo",
]

TRUSTED_DOMAINS = [
    "3cecosmetics.com",
    "smiski.com",
    "smiskifigures.com",
    "littleobsessed.com",
    "japanla.com",
    "shumistore.com",
    "msh-labo.com",
    "samurai-drugstore.jp",
    "suzykirei.com",
    "dodoskin.com",
    "fwee.us",
    "ulta.com",
    "hwahae.com",
    "wcosmetics.com.au",
    "japanmart.co.nz",
    "judydoll.com",
    "millefee.com",
    "joybeautyhub.shop",
    "intoyoucosmetics.com",
    "yesstyle.com",
    "asianbeautywholesale.com",
    "uniquebunny.com",
    "kiseki.ca",
    "oliveyoung.com",
    "stylevana.com",
]

LOVE_LINER_CREAM_FIT_PENCIL_R = {
    "source_url": (
        "https://www.msh-labo.com/c/make-up/eyeliner/1112\n"
        "https://onlineshop.japanmart.co.nz/products/love-liner-cream-fit-pencil-r-slimoval-mbr-eye-liner-0-05g\n"
        "https://www.hwahae.com/en/products/LOVELiner-Cream-Fit-Pencil-Liner-Ultra-Slim-Medium-Brown/2183519/ingredients\n"
        "https://www.yesstyle.com/en/msh-love-liner-cream-fit-pencil-medium-brown-pokemon-limited-edition/info.html/pid.1107147858\n"
        "https://wcosmetics.com.au/products/love-liner-cream-fit-pencil-r"
    ),
    "net weight": "Net. 0.05 g",
    "manufacturer": "msh Inc.",
    "coo": "Made In Japan / Fabrique au Japon",
    "source_direction": "Extend the pencil by about 1 to 2 mm and apply along the lash line. Close the cap firmly after use.",
    "ingredients": (
        "Trimethylsiloxysilicate, Methyl Trimethicone, Candelilla Wax Hydrocarbon, "
        "Hydrogenated Polyisobutene, Synthetic Wax, Acrylates/Stearyl Acrylate/Dimethicone "
        "Methacrylate Copolymer, Behenyl Alcohol, Tri(Behenic Acid/Isostearic "
        "Acid/Eicosanedioic Acid) Glyceryl, Ricinus Communis (Castor) Seed Oil, Squalane, "
        "Butyrospermum Parkii (Shea) Butter, Rosa Canina Fruit Oil, Persea Gratissima "
        "(Avocado) Oil, Camellia Japonica Seed Oil, Argania Spinosa Kernel Oil, Simmondsia "
        "Chinensis (Jojoba) Seed Oil, Macadamia Ternifolia Seed Oil, Prinsepia Utilis Seed "
        "Oil, Limnanthes Alba (Meadowfoam) Seed Oil, Carthamus Tinctorius (Safflower) Seed "
        "Oil, Helianthus Annuus (Sunflower) Seed Oil, Rosa Damascena Flower Extract, "
        "Lavandula Angustifolia (Lavender) Flower Extract, Tocopherol, Tocopheryl Acetate, "
        "Iron Oxides, Mica, Titanium Dioxide"
    ),
}

LOVE_LINER_LIQUID_EYELINER_R5 = {
    "source_url": (
        "https://www.msh-labo.com/c/loveliner/1160\n"
        "https://www.samurai-drugstore.jp/default/4570159423723.html\n"
        "https://suzykirei.com/products/381753\n"
        "https://www.dodoskin.com/products/love-liner-liquid-eyeliner-r5-0-55ml-6-colors-ultra-slim-2-types"
    ),
    "net weight": "Net. 0.55 mL",
    "manufacturer": "msh Inc.",
    "coo": "Made In Japan / Fabriqué au Japon",
    "source_direction": "Shake gently before use. Glide the tip along the lash line and close the cap firmly after use.",
    "ingredients": (
        "Water, Butylene Glycol, Acrylates Copolymer, Pentylene Glycol, Ammonium "
        "Styrene/Acrylates Copolymer, 1,2-Hexanediol, Panthenol, Glycerin, Swertia "
        "Japonica Extract, Malus Domestica Fruit Cell Culture Extract, Xanthan Gum, "
        "Rosa Centifolia Flower Extract, Sodium Hyaluronate, Lecithin, Pinus Sylvestris "
        "Cone Extract, Glycine, Sodium Metabisulfite, Camellia Sinensis Leaf Extract, "
        "Zinc Chloride, Oligopeptide-20, Oligopeptide-41, Ammonium Acrylates Copolymer, "
        "Polyglyceryl-10 Myristate, Trideceth-6 Phosphate, DPG, Glyceryl Caprylate, "
        "Caprylyl Glycol, Sodium Dehydroacetate, Phenoxyethanol, Charcoal"
    ),
}

FWEE_REFERENCE_SOURCE_RULES = [
    (
        ("fwee", "glowy", "jelly", "pot"),
        (
            "https://fwee.us/products/jelly-pot\n"
            "https://www.ulta.com/p/lip-cheek-glowy-jelly-pot-pimprod2053366"
        ),
    ),
    (
        ("fwee", "rose", "obsession", "stay", "fit", "lip", "tint"),
        (
            "https://www.yesstyle.com/en/fwee-rose-obsession/info.html/pid.1136684590\n"
            "https://www.hwahae.com/en/products/fwee-Rose-Obsession-Stay-fit-Lip-Tint-GW02-Spring-Rose/2179463"
        ),
    ),
    (
        ("fwee", "3d", "voluming", "gloss"),
        (
            "https://fwee.us/products/3d-voluming-gloss\n"
            "https://www.ulta.com/p/3d-voluming-gloss-70-pimprod2053324"
        ),
    ),
    (
        ("fwee", "pocket", "eye", "palette"),
        (
            "https://fwee.us/products/pocket-eye-palette-1\n"
            "https://www.ulta.com/p/pocket-eyeshadow-palette-pimprod2053338"
        ),
    ),
    (
        ("fwee", "oversized", "silicone", "jumbo", "brush"),
        "https://fwee.us/products/jumbo-silicone-jumbo-makeup-applicator",
    ),
    (
        ("fwee", "jumbo", "silicone", "applicator"),
        "https://fwee.us/products/jumbo-silicone-jumbo-makeup-applicator",
    ),
]

SEARCH_LANGUAGE_HINTS = [
    "",
    "English",
    "official",
    "ingredients",
    "net weight",
    "how to use",
    "全成分",
    "成分",
    "内容量",
    "容量",
    "使用方法",
    "使い方",
    "전성분",
    "성분",
    "용량",
    "사용법",
]

INGREDIENT_HEADING_RE = r"(?:major\s+ingredients?|ingredients?|ingr[eé]dients?|全成分|成分|配合成分|전성분|성분)"
NET_WEIGHT_HEADING_RE = r"(?:net\s*weight|contents?|capacity|size|内容量|容量|용량|중량)"
USAGE_HEADING_RE = r"(?:how\s+to\s+use|directions?|usage|使用方法|使い方|ご使用方法|사용법|사용방법)"
CAUTION_HEADING_RE = r"(?:cautions?|warnings?|注意事項|ご注意|사용\s*시\s*주의사항|주의사항)"

PRODUCT_ABBREVIATIONS = {
    "MBR": "Medium Brown",
    "MDBR": "Medium Brown",
    "DBR": "Dark Brown",
    "ABR": "Ash Brown",
    "BR": "Brown",
    "BK": "Black",
    "BLK": "Black",
    "RD": "Red",
    "PK": "Pink",
    "CR": "Coral",
    "OR": "Orange",
    "BE": "Beige",
    "GY": "Gray",
    "GRY": "Gray",
    "WT": "White",
}

PRODUCT_PHRASE_ALIASES = {
    "mini glow figures": "Mini Glow in the Dark Figure Blind Box",
    "mini glow figure": "Mini Glow in the Dark Figure Blind Box",
    "glow figures": "Glow in the Dark Figure",
    "glow figure": "Glow in the Dark Figure",
    "mini shero lip mud": "Shero Super Matte Lip Cheek Mud",
    "shero soft matte lip mud": "Shero Super Matte Lip Cheek Mud",
    "shero lip mud": "Shero Super Matte Lip Cheek Mud",
    "shero super matte lip mud": "Shero Super Matte Lip Cheek Mud",
    "super matte lip mud": "Shero Super Matte Lip Cheek Mud",
    "slim lip velvet": "Slim Velvet Lip Color",
    "slim velvet lip": "Slim Velvet Lip Color",
    "slim velvet lip color": "Slim Velvet Lip Color",
    "lip clolor": "Lip Color",
    "auq fit": "Aqua Fit",
    "creamfit": "Cream Fit",
    "cream fit": "Cream Fit",
    "slimoval": "Ultra Slim",
    "slim oval": "Ultra Slim",
    "ultraslim": "Ultra Slim",
    "liquideyeliner": "Liquid Eyeliner",
    "liquid eyeliner": "Liquid Eyeliner",
    "eye liner": "Eyeliner",
    "lipcheek": "Lip Cheek",
    "lip&cheek": "Lip Cheek",
}

COUNT_ITEM_TERMS = [
    "figure",
    "figures",
    "figurine",
    "figurines",
    "blind box",
    "mystery box",
    "random style",
    "random figure",
    "collectible",
    "toy",
    "brush",
    "applicator",
    "puff",
    "sponge",
    "lash curler",
    "eyelash curler",
    "tweezer",
    "sharpener",
    "mirror",
    "case",
    "pouch",
    "keyring",
    "keychain",
    "charm",
]

COLLECTIBLE_TERMS = [
    "figure",
    "figures",
    "figurine",
    "figurines",
    "blind box",
    "mystery box",
    "random style",
    "random figure",
    "collectible",
    "toy",
    "smiski",
]

TOOL_ITEM_TERMS = [
    "brush",
    "applicator",
    "puff",
    "sponge",
    "lash curler",
    "eyelash curler",
    "tweezer",
    "sharpener",
    "mirror",
]

BRAND_ALIAS_RULES = [
    {
        "keywords": ("abib",),
        "patterns": [r"\babib\b"],
        "aliases": ["Abib"],
        "domains": ["en.abib.com", "yesstyle.com", "oliveyoung.com", "hwahae.com"],
    },
    {
        "keywords": ("andbe",),
        "patterns": [r"\b&\s*be\b", r"\band\s*be\b", r"\b＆\s*be\b", r"\bbe\b"],
        "aliases": ["&be", "andbe"],
        "domains": ["andbe-official.com", "yesstyle.com", "japaniful.com", "hwahae.com"],
    },
    {
        "keywords": ("3ce",),
        "patterns": [r"\b3\s*ce\b", r"\b3ce\b", r"\b3\s*concept\s*eyes\b"],
        "aliases": ["3CE", "3 Concept Eyes"],
        "domains": ["3cecosmetics.com", "yesstyle.com", "asianbeautywholesale.com", "hwahae.com", "stylevana.com"],
    },
    {
        "keywords": ("smiski",),
        "patterns": [r"\bsmiski\b", r"\bdreams\s*smiski\b"],
        "aliases": ["Smiski", "Dreams Smiski", "SMISKI"],
        "domains": ["smiski.com", "smiskifigures.com", "littleobsessed.com", "japanla.com", "shumistore.com"],
    },
    {
        "keywords": ("love", "liner"),
        "patterns": [
            r"\bmsh\s*[- ]*\s*love\s*liner\b",
            r"\bloveer\s*liner\b",
            r"\blove\s*liner\b",
            r"\bloveliner\b",
        ],
        "aliases": ["MSH Love Liner", "Love Liner", "LoveLiner"],
        "domains": [
            "msh-labo.com",
            "yesstyle.com",
            "oliveyoung.com",
            "samurai-drugstore.jp",
            "suzykirei.com",
            "dodoskin.com",
        ],
    },
    {
        "keywords": ("judydoll",),
        "patterns": [r"\bjudy\s*doll\b", r"\bjudydoll\b"],
        "aliases": ["Judydoll", "JudyDoll", "Judy Doll"],
        "domains": ["judydoll.com", "yesstyle.com", "oliveyoung.com", "kiseki.ca"],
    },
    {
        "keywords": ("into", "you"),
        "patterns": [r"\binto\s*you\b", r"\bity\b"],
        "aliases": ["INTO YOU", "INTO YOU Cosmetics", "ITY"],
        "domains": ["intoyoucosmetics.com", "yesstyle.com", "asianbeautywholesale.com", "uniquebunny.com", "kiseki.ca"],
    },
    {
        "keywords": ("fwee",),
        "patterns": [r"\bfwee\b"],
        "aliases": ["fwee", "FWEE"],
        "domains": ["fwee.us", "yesstyle.com", "oliveyoung.com", "ulta.com", "hwahae.com"],
    },
    {
        "keywords": ("millefee",),
        "patterns": [r"\bmille\s*fee\b", r"\bmillefee\b"],
        "aliases": ["MilleFee", "Mille Fee"],
        "domains": ["millefee.com", "yesstyle.com", "oliveyoung.com"],
    },
]

PRODUCT_NAME_FR_PHRASES = {
    "Mini Glow in the Dark Figure Blind Box": "mini figurine phosphorescente en boîte surprise",
    "Mini Glow Figure Blind Box": "mini figurine phosphorescente en boîte surprise",
    "Mini Glow Figures": "mini figurines phosphorescentes",
    "Mini Glow Figure": "mini figurine phosphorescente",
    "Glow in the Dark Figure": "figurine phosphorescente",
    "Glow in the Dark Figures": "figurines phosphorescentes",
    "Bath Series": "série bain",
    "Blind Box": "boîte surprise",
    "Mystery Box": "boîte mystère",
    "Random Style": "modèle aléatoire",
    "Airy Lip Cheek Mud": "baume mat aérien lèvres et joues",
    "Airy Lip & Cheek Mud": "baume mat aérien lèvres et joues",
    "Shero Super Matte Lip Cheek Mud": "baume mat Shero lèvres et joues",
    "Shero Super Matte Lip Mud": "baume mat Shero lèvres",
    "Mini Shero Lip Mud": "mini baume mat Shero lèvres",
    "Lip Cheek Glowy Jelly Pot": "pot gelée éclat lèvres et joues",
    "Lip&Cheek Glowy Jelly Pot": "pot gelée éclat lèvres et joues",
    "Lip & Cheek Glowy Jelly Pot": "pot gelée éclat lèvres et joues",
    "Lip Cheek Mud": "baume mat lèvres et joues",
    "Lip & Cheek Mud": "baume mat lèvres et joues",
    "Liquid Blush Serum": "sérum blush liquide",
    "Six-color Blush Palette": "palette de fards à joues six couleurs",
    "Six-Color Blush Palette": "palette de fards à joues six couleurs",
    "Idol Highlighter Palette": "palette illuminateur Idol",
    "Highlighter Palette": "palette illuminateur",
    "Pocket Eye Palette": "palette pour les yeux compacte",
    "Eye Palette": "palette pour les yeux",
    "Eyeshadow Palette": "palette de fards à paupières",
    "Eye Shadow Palette": "palette de fards à paupières",
    "3D Voluming Gloss": "gloss volumisant 3D",
    "Voluming Gloss": "gloss volumisant",
    "Mood Recipe Matte Lip Color": "rouge à lèvres mat Mood Recipe",
    "Matte Lip Color": "rouge à lèvres mat",
    "Lip Color": "rouge à lèvres",
    "Stay-Fit Lip Tint": "teinte à lèvres tenue longue durée",
    "Stay Fit Lip Tint": "teinte à lèvres tenue longue durée",
    "Cushion Foundation": "fond de teint coussin",
    "Sheet Mask": "masque en feuille",
    "Mild Acidic pH Sheet Mask": "masque en feuille pH doux acide",
    "Rose Obsession": "obsession rose",
    "Glow Lipstick": "rouge à lèvres éclat",
    "Glowing Lipstick": "rouge à lèvres éclat",
    "Liquid Eyeliner": "traceur liquide pour les yeux",
    "Liquid Eye Liner": "traceur liquide pour les yeux",
    "Cream Fit Pencil": "crayon pour les yeux Cream Fit",
    "CreamfIT Pencil": "crayon pour les yeux Cream Fit",
    "Pencil Liner": "crayon pour les yeux",
    "Eye Liner": "traceur pour les yeux",
    "Eyeliner": "traceur pour les yeux",
    "Lipstick": "rouge à lèvres",
    "Lip Tint": "teinte à lèvres",
    "Blush Serum": "sérum blush",
    "Blush Palette": "palette de fards à joues",
    "Blush": "blush",
    "Ultra Slim": "ultra fin",
    "SlimOval": "ovale ultra fin",
    "Slim Oval": "ovale ultra fin",
    "Lash Curler": "recourbe-cils",
    "Eyelash Curler": "recourbe-cils",
    "Silicone Brush": "pinceau en silicone",
    "Jumbo Brush": "pinceau jumbo",
    "Brush": "pinceau",
    "Applicator": "applicateur",
    "Puff": "houppette",
    "Sponge": "éponge",
    "Tweezer": "pince à épiler",
    "Sharpener": "taille-crayon",
    "Mirror": "miroir",
    "PENDANT KEYRING": "porte-clés pendentif",
    "KEYRING": "porte-clés",
    "Figure": "figurine",
    "Figures": "figurines",
    "Glow": "phosphorescent",
    "Bath": "bain",
    "Mini": "mini",
    "Lip": "lèvres",
    "Cheek": "joues",
}

SHADE_FR_TERMS = {
    "Medium Brown": "brun moyen",
    "Ash Brown": "brun cendré",
    "Red Brown": "brun rouge",
    "Rose Pink": "rose rosé",
    "Ice Blue": "bleu glacé",
    "Lavender Whip": "crème lavande",
    "Milky": "laiteux",
    "Custard": "crème anglaise",
    "Honey Peach": "pêche au miel",
    "Mellow Mango": "mangue douce",
    "Juicy Punch": "punch juteux",
    "Sugar Coat": "enrobage sucré",
    "Lolly": "sucette",
    "Sweets": "confiseries",
    "Squeezed": "pressé",
    "Sour Gummy": "bonbon gomme acide",
    "Cherry Ppo": "cerise",
    "Berry Jam": "confiture de baies",
    "Cream Tea": "thé à la crème",
    "Rosy Glaze": "glaçage rosé",
    "Caramelized": "caramélisé",
    "Sugar Powder": "poudre de sucre",
    "Spring Rose": "rose printemps",
    "Coco Rose": "rose coco",
    "Sweet Rose": "rose douce",
    "Apple Rose": "rose pomme",
    "Fresh Rose": "rose fraîche",
    "Plum Rose": "rose prune",
    "Burnt Rose": "rose brûlée",
    "Cinnamon Rose": "rose cannelle",
    "Peach Rose": "rose pêche",
    "Dried Rose": "rose séchée",
    "Dusty Rose": "rose poussiéreuse",
    "Pink Milk": "lait rose",
    "Rosewood": "bois de rose",
    "Mauve Taupe": "taupe mauve",
    "Yogurt Peach": "pêche yogourt",
    "Strawberry Whip": "crème fouettée à la fraise",
    "Salty Caramel": "caramel salé",
    "Grape Bonbon": "bonbon raisin",
    "Cherry Cola": "cola cerise",
    "Dirty Cola": "cola foncé",
    "Fig": "figue",
    "Salmon": "saumon",
    "Black": "noir",
    "Brown": "brun",
    "Red": "rouge",
    "Pink": "rose",
    "Peach": "pêche",
    "Coral": "corail",
    "Mango": "mangue",
    "Lavender": "lavande",
}

KNOWN_ONLINE_PRODUCTS = {
    "1129343972": {
        "source_url": "https://www.asianbeautywholesale.com/en/into-you-glowing-lipstick-8-colors-gl08-red-brown-3g/info.html/pid.1129343972\nhttps://www.intoyoucosmetics.com/en-ca/products/into-you-glow-lipstick?variant=57958599524655\nhttps://www.intoyoucosmetics.com/en-ca/products/into-you-glow-lipstick?variant=57958599590191\nhttps://www.yesstyle.com/en/into-you-glowing-lipstick-8-colors-gl08-red-brown-3g/info.html/pid.1129343972\nhttps://www.uniquebunny.com/products/into-you-glow-lipstick\nhttps://www.intoyoucosmetics.com/en-ca/pages/about-us",
        "net weight": "Net. 3 g",
        "source_direction": "Apply a small amount directly to lips. Store below 25°C and refrigerate if the product softens.",
        "ingredients": "Isostearyl Isostearate, Polyglyceryl-2 Triisostearate, Diisostearyl Malate, Sorbitan Isostearate, Paraffin, Trimethylpentaphenyl Trisiloxane, Microcrystalline Wax, Pentaerythrityl Isostearate, Euphorbia Cerifera (Candelilla) Wax, 1,2-Hexanediol, PEG/PPG-10/1 Dimethicone, CI 77891, CI 19140, CI 45410, CI 77491, Fragrance, CI 77499, Pentaerythrityl Tetraisostearate",
        "manufacturer": "HONGKONG LETS INTERNATIONAL TRADING LIMITED",
    },
    "1126245093": {
        "source_url": "https://www.intoyoucosmetics.com/en-gb/products/airy-lip-cheek-mud\nhttps://www.uniquebunny.com/fr/products/into-you-airy-lip-cheek-mud\nhttps://www.uniquebunny.com/products/into-you-airy-lip-cheek-mud\nhttps://www.yesstyle.com/en/into-you-airy-lip-cheek-mud-5-colors-c1-c5-c5-mauve-taupe-1-8g/info.html/pid.1126244966",
        "net weight": "Net. 2 g",
        "source_direction": "Apply a proper amount evenly on lips, or dab onto cheeks and blend with fingertips.",
        "coo": "Made In China / Fabriqué En Chine",
        "ingredients_fr": "Diméthicone, polymère croisé de diméthicone, triisostéarate de polyglycéryle-2, cyclopentasiloxane, cyclohexasiloxane, CI 19140, 1,2-hexanediol, silylate de silice, cétyl PEG/PPG-10/1 diméthicone, talc, CI 77491, CI 77891, extrait de Melaleuca Alternifolia, huile d'onagre, acétate d'alpha-tocophéryle, C30-45 Alkyl Diméthicone",
    },
    "1137202898": {
        "source_url": "https://www.yesstyle.com/en/into-you-six-color-blush-palette-six-color-blush-palette-15g/info.html/pid.1137202898",
        "net weight": "Net. 15 g",
        "source_direction": "Apply the blush to cheeks with a brush. Mix, match, and layer shades as desired.",
        "ingredients": "Talc, Mica, CI 77891, Synthetic Fluorphlogopite, Silica, Magnesium Myristate, Vinyl Dimethicone/Methicone Silsesquioxane Crosspolymer, Octyldodecanol, Isostearyl Neopentanoate, Dimethicone, Methyl Methacrylate Crosspolymer, Aluminum Myristate, Diisostearyl Malate, CI 77491, CI 75470, Ethylhexylglycerin, CI 77492, Glyceryl Caprylate, Lauroyl Lysine, CI 77499, Dimethicone Crosspolymer, Triethoxycaprylylsilane, Cocos Nucifera (Coconut) Oil, Alumina, Water, Aluminum Hydroxide, CI 73360, CI 19140, CI 45410, CI 77007, Boron Nitride",
        "manufacturer": "HONGKONG LETS INTERNATIONAL TRADING LIMITED",
    },
}

JUDYDOLL_LIQUID_BLUSH_SERUM = {
    "source_url": (
        "https://www.yesstyle.com/en/judydoll-liquid-blush-serum-4-colors-01-fig-5g/info.html/pid.1136648925\n"
        "https://judydoll.com/products/liquid-blush-serum"
    ),
    "net weight": "Net. 5 g",
    "coo": "Made In China / Fabriqué En Chine",
    "source_direction": "Lightly dab after foundation and before setting powder, then blend evenly across cheeks.",
    "shades": {
        "01-fig": "water, CI 77891, cyclopentasiloxane, isododecane, diisopropyl sebacate, butylene glycol, lauryl PEG-10 tri(trimethylsiloxy) silane, 1,2-pentanediol, diphenylmethoxy silane, trimethylsilyl silicate, CI 77019, polydimethylsiloxane, cyclohexasiloxane, squalane, cetearyl PEG/PPG-10/1 polydimethylsiloxane, polymethyl methacrylate, distearyldimonium lithium montmorillonite, magnesium sulfate, phenoxyethanol, polydimethylsiloxane/vinylpolydimethylsiloxane crosspolymer, polyhydroxystearic acid, HDI/trihydroxymethyl hexyl lactone crosspolymer, CI 77491, aluminum hydroxide, triethoxy octylsilane, CI 77492, polydimethylsiloxane crosspolymer, ethylhexylglycerin, vinylpolydimethylsiloxane/polymethylsilsesquioxane crosspolymer, disodium ethylenediaminetetraacetate, CI 73360, CI 77499, silica, glycerin, 1,3-propanediol, tocopherol (vitamin E), pentaerythrityl tetra(butylated hydroxytoluene) ester, collagen, PPG-13-decyloleate-24, 1,2-hexanediol, acetyl hexapeptide-8, palmitoyl pentapeptide-4",
        "02-pink-milk": "water, CI 77891, cyclopentasiloxane, isododecane, diisopropyl sebacate, butylene glycol, lauryl PEG-10 tri(trimethylsiloxy) silane, 1,2-pentanediol, diphenylmethoxy silane, trimethylsilyl silicate, polydimethylsiloxane, cyclohexasiloxane, squalane, cetearyl PEG/PPG-10/1 polydimethylsiloxane, polymethyl methacrylate, distearyldimonium lithium montmorillonite, magnesium sulfate, phenoxyethanol, CI 77019, polydimethylsiloxane/vinylpolydimethylsiloxane crosspolymer, polyhydroxystearic acid, HDI/trihydroxymethyl hexyl lactone crosspolymer, CI 77492, aluminum hydroxide, CI 77007, triethoxy octylsilane, polydimethylsiloxane crosspolymer, ethylhexylglycerin, vinylpolydimethylsiloxane/polymethylsilsesquioxane crosspolymer, CI 73360, CI 77499, disodium ethylenediaminetetraacetate, silica, glycerin, 1,3-propanediol, tocopherol (vitamin E), pentaerythrityl tetra(butylated hydroxytoluene) ester, collagen, PPG-13-decyloleate-24, 1,2-hexanediol, acetyl hexapeptide-8, palmitoyl pentapeptide-4",
        "03-salmon": "water, CI 77891, cyclopentasiloxane, isododecane, diisopropyl sebacate, butylene glycol, lauryl PEG-10 tri(trimethylsiloxy) silane, 1,2-pentanediol, diphenylmethoxy silane, trimethylsilyl silicate, polydimethylsiloxane, CI 77019, cyclohexasiloxane, squalane, cetearyl PEG/PPG-10/1 polydimethylsiloxane, polymethyl methacrylate, CI 77491, distearyldimonium lithium montmorillonite, magnesium sulfate, phenoxyethanol, polydimethylsiloxane/vinylpolydimethylsiloxane crosspolymer, polyhydroxystearic acid, HDI/trihydroxymethyl hexyl lactone crosspolymer, CI 77492, aluminum hydroxide, triethoxy octylsilane, polydimethylsiloxane crosspolymer, ethylhexylglycerin, vinylpolydimethylsiloxane/polymethylsilsesquioxane crosspolymer, CI 77499, CI 73360, disodium ethylenediaminetetraacetate, silica, glycerin, 1,3-propanediol, tocopherol (vitamin E), pentaerythrityl tetra(butylated hydroxytoluene) ester, collagen, PPG-13-decyloleate-24, 1,2-hexanediol, acetyl hexapeptide-8, palmitoyl pentapeptide-4",
        "04-rosewood": "water, cyclopentasiloxane, CI 77019, isododecane, diisopropyl sebacate, CI 77891, butylene glycol, lauryl PEG-10 tri(trimethylsiloxy) silane, 1,2-pentanediol, diphenylmethoxy silane, trimethylsilyl silicate, polydimethylsiloxane, cyclohexasiloxane, squalane, cetearyl PEG/PPG-10/1 polydimethylsiloxane, polymethyl methacrylate, distearyldimonium lithium montmorillonite, magnesium sulfate, phenoxyethanol, polydimethylsiloxane/vinylpolydimethylsiloxane crosspolymer, polyhydroxystearic acid, HDI/trihydroxymethyl hexyl lactone crosspolymer, CI 77491, triethoxy octylsilane, CI 77492, CI 73360, CI 77007, CI 77499, aluminum hydroxide, polydimethylsiloxane crosspolymer, ethylhexylglycerin, vinylpolydimethylsiloxane/polymethylsilsesquioxane crosspolymer, disodium ethylenediaminetetraacetate, silica, glycerin, 1,3-propanediol, tocopherol (vitamin E), pentaerythrityl tetra(butylated hydroxytoluene) ester, collagen, PPG-13-decyloleate-24, 1,2-hexanediol, acetyl hexapeptide-8, palmitoyl pentapeptide-4",
    },
}

MILLEFEE_IDOL_HIGHLIGHTER_PALETTE = {
    "net weight": "Net. 11 g",
    "coo": "Made In China / Fabriqué En Chine",
    "source_direction": "Apply to high points of the face with a brush. Layer as desired for added glow.",
    "shades": {
        "01-ice-blue": {
            "source_url": (
                "https://www.yesstyle.com/en/millefee-idol-highlighter-palette-01-ice-blue/info.html/pid.1137196647\n"
                "https://millefee.com/products/idol-highlighter-palette"
            ),
            "ingredients": "Synthetic Fluorphlogopite, Titanium Dioxide, Dimethicone, Silica, Phenyl Trimethicone, Calcium Aluminum Borosilicate, Diisostearyl Malate, Petrolatum, Isononyl Isononanoate, Synthetic Wax, Polyisobutene, Dimethicone/Vinyl Dimethicone Crosspolymer, Trimethylsiloxysilicate, Methyl Methacrylate Crosspolymer, Tin Oxide, Microcrystalline Wax, Caprylyl Glycol, Hydroxystearic Acid, Polymethylsilsesquioxane, Iron Oxides, Dimethicone Crosspolymer, Ethylhexylglycerin, Triethoxycaprylylsilane, Stearic Acid, Red 226, Barium Sulfate, Yellow 4, Mica, Blue 1",
        },
        "02-rose-pink": {
            "source_url": (
                "https://www.yesstyle.com/en/millefee-idol-highlighter-palette-02-rose-pink/info.html/pid.1137196649\n"
                "https://millefee.com/products/idol-highlighter-palette"
            ),
            "ingredients": "Synthetic Fluorphlogopite, Titanium Dioxide, Isononyl Isononanoate, Silica, Dimethicone, Polyisobutene, Phenyl Trimethicone, Calcium Aluminum Borosilicate, Polymethylsilsesquioxane, Dimethicone Crosspolymer, Hydroxystearic Acid, Microcrystalline Wax, Caprylyl Glycol, Tin Oxide, Triethoxycaprylylsilane, Ethylhexylglycerin, Stearic Acid, Iron Oxides, Red 226, Barium Sulfate, Blue 1",
        },
    },
    "barcodes": {
        "1137196647": "01-ice-blue",
        "1137196649": "02-rose-pink",
    },
}


def ensure_dirs() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    EXPORT_DIR.mkdir(exist_ok=True)


def password_hash(password: str, salt: str | None = None) -> str:
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), salt.encode("utf-8"), 180_000
    ).hex()
    return f"{salt}${digest}"


def verify_password(password: str, stored: str) -> bool:
    if "$" not in stored:
        return False
    salt, expected = stored.split("$", 1)
    return secrets.compare_digest(password_hash(password, salt).split("$", 1)[1], expected)


def db() -> sqlite3.Connection:
    ensure_dirs()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('admin', 'viewer')),
            active INTEGER NOT NULL DEFAULT 1
        )
        """
    )
    conn.execute(
        """
        INSERT OR IGNORE INTO users(username, password_hash, role, active)
        VALUES(?,?,?,1)
        """,
        ("admin", password_hash("change-me-now"), "admin"),
    )
    conn.commit()
    return conn


def authenticate(username: str, password: str) -> dict[str, Any] | None:
    with db() as conn:
        row = conn.execute(
            "SELECT username, password_hash, role, active FROM users WHERE username=?",
            (username.strip(),),
        ).fetchone()
    if not row or not row["active"] or not verify_password(password, row["password_hash"]):
        return None
    return {"username": row["username"], "role": row["role"]}


def manage_users() -> None:
    st.subheader("Users")
    with db() as conn:
        rows = conn.execute(
            "SELECT username, role, active FROM users ORDER BY username"
        ).fetchall()
    users = pd.DataFrame([dict(r) for r in rows])
    st.dataframe(users, use_container_width=True, hide_index=True)

    with st.form("add_user"):
        st.caption("Add approved user")
        col1, col2, col3 = st.columns([2, 2, 1])
        username = col1.text_input("Username")
        password = col2.text_input("Temporary password", type="password")
        role = col3.selectbox("Role", ["viewer", "admin"])
        if st.form_submit_button("Add user"):
            if not username or not password:
                st.error("Username and password are required.")
            else:
                try:
                    with db() as conn:
                        conn.execute(
                            "INSERT INTO users(username, password_hash, role, active) "
                            "VALUES(?,?,?,1)",
                            (username.strip(), password_hash(password), role),
                        )
                        conn.commit()
                    st.success(f"Added {username}.")
                except sqlite3.IntegrityError:
                    st.error("That username already exists.")

    with st.form("update_user"):
        st.caption("Update approved user")
        selected = st.selectbox("User", users["username"].tolist() if not users.empty else [])
        col1, col2, col3 = st.columns(3)
        new_role = col1.selectbox("New role", ["viewer", "admin"])
        active = col2.checkbox("Active", value=True)
        new_password = col3.text_input("New password", type="password")
        if st.form_submit_button("Save user"):
            with db() as conn:
                conn.execute(
                    "UPDATE users SET role=?, active=? WHERE username=?",
                    (new_role, int(active), selected),
                )
                if new_password:
                    conn.execute(
                        "UPDATE users SET password_hash=? WHERE username=?",
                        (password_hash(new_password), selected),
                    )
                conn.commit()
            st.success("User updated.")


class LinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[tuple[str, str]] = []
        self._href: str | None = None
        self._text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "a":
            attrs_dict = dict(attrs)
            href = attrs_dict.get("href")
            if href and (href.startswith("http") or href.startswith("/url?")):
                self._href = href
                self._text = []

    def handle_data(self, data: str) -> None:
        if self._href:
            self._text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == "a" and self._href:
            text = html.unescape(" ".join(self._text)).strip()
            if text and len(text) > 8:
                self.links.append((text, self._href))
            self._href = None
            self._text = []


@st.cache_data(ttl=60 * 60)
def search_web(query: str) -> list[dict[str, str]]:
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9,fr;q=0.8",
    }
    urls = [
        f"https://www.google.com/search?q={quote_plus(query)}",
        f"https://duckduckgo.com/html/?q={quote_plus(query)}",
        f"https://www.bing.com/search?q={quote_plus(query)}",
    ]
    results: list[dict[str, str]] = []
    for url in urls:
        try:
            response = requests.get(url, timeout=12, headers=headers)
            response.raise_for_status()
        except requests.RequestException:
            continue
        parser = LinkParser()
        parser.feed(response.text)
        for title, href in parser.links:
            domain = urlparse(href).netloc.lower()
            if any(bad in domain for bad in ["facebook", "instagram", "tiktok", "pinterest"]):
                continue
            results.append({"title": title, "url": href})
        if results:
            break
    deduped: list[dict[str, str]] = []
    seen = set()
    for result in results:
        result["url"] = clean_search_url(result["url"])
        if is_noise_url(result["url"]):
            continue
        key = result["url"].split("&")[0]
        if key not in seen:
            seen.add(key)
            deduped.append(result)
    return sorted(deduped, key=lambda item: source_rank(item["url"]))[:25]


def clean_search_url(url: str) -> str:
    parsed = urlparse(url)
    if parsed.path == "/url":
        target = parse_qs(parsed.query).get("q", [""])[0]
        if target.startswith("http"):
            return target
    if "bing.com" in parsed.netloc and parsed.path.startswith("/ck/"):
        target = parse_qs(parsed.query).get("u", [""])[0]
        if target.startswith("a1"):
            try:
                import base64

                decoded = base64.urlsafe_b64decode(target[2:] + "==").decode("utf-8", "ignore")
                if decoded.startswith("http"):
                    return decoded
            except Exception:
                pass
        if target.startswith("http"):
            return unquote(target)
    keep_params = {}
    for key, values in parse_qs(parsed.query).items():
        if key.lower() in {"variant", "pid"}:
            keep_params[key] = values[-1]
    clean_query = urlencode(keep_params)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", clean_query, ""))


def is_noise_url(url: str) -> bool:
    parsed = urlparse(url)
    domain = parsed.netloc.lower()
    path = parsed.path.lower()
    if any(bad in domain for bad in ["facebook", "instagram", "tiktok", "pinterest"]):
        return True
    noisy_parts = [
        "/list.html",
        "/brand/",
        "/brands/",
        "/collections/",
        "/collection/",
        "/search",
        "/tag/",
        "/blog",
        "/blogs/",
        "/category/",
        "/categories/",
        "/cart",
        "/account",
    ]
    if any(part in path for part in noisy_parts):
        if "/products/" not in path and "/info.html" not in path:
            return True
    return False


def source_rank(url: str) -> int:
    domain = domain_from_url(url)
    for idx, trusted in enumerate(TRUSTED_DOMAINS):
        if trusted in domain:
            return idx
    if any(market in domain for market in ["amazon.", "ebay.", "temu.", "aliexpress."]):
        return 180
    if any(noisy in domain for noisy in ["baidu.", "wikipedia.", "reddit."]):
        return 220
    return 100


def product_detail_bonus(url: str) -> float:
    parsed = urlparse(url)
    domain = parsed.netloc.lower()
    path = parsed.path.lower()
    if "yesstyle.com" in domain and "/info.html" in path and "/pid." in path:
        return 9.0
    if "3cecosmetics.com" in domain and "/all-products/" in path:
        return 7.0
    if "/products/" in path:
        return 6.5
    if path.endswith(".html") and not is_noise_url(url):
        return 4.0
    return 0.0


def exact_product_url(url: str, product: str) -> bool:
    if is_noise_url(url):
        return False
    tokens = []
    for alias in expanded_product_names(product):
        tokens.extend(search_tokens(alias))
    tokens = list(dict.fromkeys(tokens))
    if not tokens:
        return product_detail_bonus(url) > 0
    haystack = searchable_text(url)
    product_words = [token for token in tokens if token not in {"serum", "liquid", "color", "colors"}]
    hits = sum(1 for token in product_words if token in haystack)
    shade = shade_key(product)
    shade_ok = not shade or shade in haystack
    return product_detail_bonus(url) > 0 and hits >= max(2, min(4, len(product_words) - 1)) and shade_ok


@st.cache_data(ttl=60 * 60 * 24)
def fetch_text(url: str) -> str:
    text = fetch_direct_text(url)
    if len(text) > 200:
        return text
    text = fetch_shopify_product_text(url)
    if len(text) > 200:
        return text
    text = fetch_jina_text(url)
    if len(text) > 200:
        return text
    return ""


def fetch_direct_text(url: str) -> str:
    try:
        response = requests.get(
            url,
            timeout=8,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/124.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9,fr;q=0.8,zh-CN;q=0.7,ja;q=0.6,ko;q=0.6",
            },
        )
        response.raise_for_status()
    except requests.RequestException:
        return ""
    embedded_text = extract_embedded_product_text(response.text)
    text = re.sub(r"<script.*?</script>|<style.*?</style>", " ", response.text, flags=re.S | re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    return html.unescape(re.sub(r"\s+", " ", f"{text} {embedded_text}")).strip()


def extract_embedded_product_text(raw_html: str) -> str:
    pieces: list[str] = []
    pieces.extend(extract_escaped_content_blocks(raw_html))
    for script in re.findall(r"<script[^>]*>(.*?)</script>", raw_html, flags=re.S | re.I):
        if not re.search(r"ingredients?|product information|net weight|country of origin|how to use", script, flags=re.I):
            continue
        text = html.unescape(script)
        text = html.unescape(text)
        text = text.replace("\\n", " ").replace("\\/", "/").replace('\\"', '"')
        text = re.sub(r"<[^>]+>", " ", text)
        text = re.sub(r"[\{\}\[\]]", " ", text)
        text = re.sub(r'"(?:title|content|name|description)"\s*:\s*', " ", text, flags=re.I)
        text = re.sub(r"\\u[0-9a-fA-F]{4}", " ", text)
        pieces.append(text)
    return html.unescape(re.sub(r"\s+", " ", " ".join(pieces))).strip()


def extract_escaped_content_blocks(raw_html: str) -> list[str]:
    blocks: list[str] = []
    pattern = r"&quot;title&quot;:&quot;([^&]{3,80})&quot;,&quot;content&quot;:&quot;(.*?)&quot;\}"
    useful_titles = re.compile(
        r"ingredients?|major ingredients?|product information|description|details|how to use|directions?|"
        r"cautions?|warnings?|country of origin|manufacturer",
        flags=re.I,
    )
    for title, content in re.findall(pattern, raw_html, flags=re.S | re.I):
        if not useful_titles.search(html.unescape(title)):
            continue
        text = html.unescape(html.unescape(content))
        text = text.replace("\\n", " ").replace("\\/", "/").replace('\\"', '"')
        text = re.sub(r"<[^>]+>", " ", text)
        blocks.append(f"{html.unescape(title)}: {text}")
    return blocks


def fetch_jina_text(url: str) -> str:
    try:
        response = requests.get(
            f"https://r.jina.ai/{url}",
            timeout=12,
            headers={
                "User-Agent": "Mozilla/5.0 label-research-tool/1.0",
                "Accept": "text/plain, text/markdown, */*",
            },
        )
        response.raise_for_status()
    except requests.RequestException:
        return ""
    return html.unescape(re.sub(r"\s+", " ", response.text)).strip()


def fetch_shopify_product_text(url: str) -> str:
    parsed = urlparse(url)
    if "/products/" not in parsed.path:
        return ""
    handle = parsed.path.split("/products/", 1)[1].strip("/").split("/", 1)[0]
    if not handle:
        return ""
    product_json_url = f"{parsed.scheme}://{parsed.netloc}/products/{handle}.js"
    try:
        response = requests.get(
            product_json_url,
            timeout=8,
            headers={
                "User-Agent": "Mozilla/5.0 label-research-tool/1.0",
                "Accept": "application/json,text/plain,*/*",
            },
        )
        response.raise_for_status()
        data = response.json()
    except (requests.RequestException, ValueError):
        return ""
    pieces = [
        str(data.get("title", "")),
        str(data.get("vendor", "")),
        re.sub(r"<[^>]+>", " ", str(data.get("description", ""))),
    ]
    for variant in data.get("variants", []) or []:
        pieces.append(str(variant.get("title", "")))
        pieces.append(str(variant.get("sku", "")))
    return html.unescape(re.sub(r"\s+", " ", " ".join(pieces))).strip()


@st.cache_data(ttl=60 * 60 * 24)
def hotlist_text() -> str:
    return fetch_text(HOTLIST_URL)


@st.cache_data
def builtin_reference_data() -> dict[str, dict[str, str]]:
    if not BUILTIN_REFERENCE_PATH.exists():
        return {}
    return json.loads(BUILTIN_REFERENCE_PATH.read_text(encoding="utf-8"))


def normalized_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    headers = {}
    for cell in ws[1]:
        if cell.value:
            key = str(cell.value).strip().lower()
            headers[key] = cell.column
    return headers


def find_column(headers: dict[str, int], *needles: str) -> int | None:
    for needle in needles:
        needle_l = needle.lower()
        for header, col in headers.items():
            if needle_l == header or needle_l in header:
                return col
    return None


def copy_cell_style(src: openpyxl.cell.cell.Cell, dst: openpyxl.cell.cell.Cell) -> None:
    if src.has_style:
        dst._style = copy.copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy.copy(src.alignment)


def is_need_review_value(value: Any) -> bool:
    return str(value or "").strip().lower() == "need to review"


def has_need_review_fill(cell: openpyxl.cell.cell.Cell) -> bool:
    fill = cell.fill
    if fill.fill_type != "solid":
        return False
    rgb = str(fill.fgColor.rgb or "").upper()
    return rgb in NEED_REVIEW_FILL_RGBS


def apply_need_review_highlight(cell: openpyxl.cell.cell.Cell, clear_existing: bool = False) -> None:
    if is_need_review_value(cell.value):
        cell.fill = copy.copy(NEED_REVIEW_FILL)
    elif clear_existing and has_need_review_fill(cell):
        cell.fill = openpyxl.styles.PatternFill(fill_type=None)


def highlight_need_review_cells(wb: openpyxl.Workbook) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                apply_need_review_highlight(cell)


def remove_manufacturer_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    headers = normalized_headers(ws)
    cols = [
        col
        for header, col in headers.items()
        if header == "manufacturer" or "manufacturer" in header
    ]
    for col in sorted(set(cols), reverse=True):
        ws.delete_cols(col)


def add_audit_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    remove_manufacturer_columns(ws)
    headers = normalized_headers(ws)
    for label in ["source websites", "source notes", "row status"]:
        col = headers.get(label)
        if col:
            ws.delete_cols(col)
            headers = normalized_headers(ws)

    source_col = headers.get("source url")
    if source_col and source_col != ws.max_column:
        target_col = ws.max_column + 1
        for row_idx in range(1, ws.max_row + 1):
            src = ws.cell(row_idx, source_col)
            dst = ws.cell(row_idx, target_col)
            dst.value = src.value
            copy_cell_style(src, dst)
        ws.delete_cols(source_col)
        headers = normalized_headers(ws)

    if "source url" not in headers:
        col = ws.max_column + 1
        ws.cell(1, col).value = "Source Url"
        copy_cell_style(ws.cell(1, 1), ws.cell(1, col))
        headers = normalized_headers(ws)
    return headers


def net_weight_from_name(name: str) -> str | None:
    match = re.search(r"(\d+(?:\.\d+)?)\s*(ml|mL|ML|g|G)\s*(?:\+\s*(pendant keyring|1\s*pcs|pcs))?", name)
    if not match:
        return None
    amount, unit, extra = match.group(1), match.group(2), match.group(3)
    unit = "mL" if unit.lower() == "ml" else "g"
    result = f"Net. {amount} {unit}"
    if extra:
        result += " + 1 PCS"
    return result


def net_weight_from_text(text: str) -> str | None:
    patterns = [
        rf"{NET_WEIGHT_HEADING_RE}\s*[:：]?\s*(\d+(?:\.\d+)?)\s*(g|gram|grams|ml|mL|ML)\s*(?:/[^x]{{0,20}})?\s*[x×]\s*(\d{{1,3}})",
        r"(\d+(?:\.\d+)?)\s*(g|gram|grams|ml|mL|ML)\s*(?:/[^x]{0,20})?\s*[x×]\s*(\d{1,3})",
        rf"{NET_WEIGHT_HEADING_RE}\s*[:：]?\s*(\d+(?:\.\d+)?)\s*(g|gram|grams|ml|mL|ML)",
        r"net\s*weight\s*[:：]?\s*(\d+(?:\.\d+)?)\s*(g|gram|grams|ml|mL|ML)",
        r"(\d+(?:\.\d+)?)\s*(g|gram|grams|ml|mL|ML)\s*/\s*0\.",
        r"(\d+(?:\.\d+)?)\s*(g|gram|grams|ml|mL|ML)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if match:
            amount, unit = match.group(1), match.group(2)
            unit = "mL" if unit.lower() == "ml" else "g"
            if len(match.groups()) >= 3 and match.group(3):
                return f"Net. {amount} {unit} x {int(match.group(3))} PCS"
            return f"Net. {amount} {unit}"
    return None


def compatible_net_weight(product: str, net_weight: str | None) -> str | None:
    if not net_weight:
        return None
    clean_product = searchable_text(product)
    clean_weight = searchable_text(net_weight)
    solid_lip_terms = ["lipstick", "lip color", "lipcolour", "lip stick", "matte lip color", "velvet lip color"]
    if "ml" in clean_weight and any(term in clean_product for term in solid_lip_terms):
        if not any(term in clean_product for term in ["liquid", "tint", "gloss", "serum", "oil"]):
            return None
    return net_weight


def pcs_count_from_text(product: str, text: str = "") -> str | None:
    source = searchable_text(f"{product} {text}")
    match = re.search(r"\b(\d{1,2})\s*(?:pcs|pc|pieces|piece|packs|pack|count|ct)\b", source)
    if match:
        return f"Net. {int(match.group(1))} PCS"
    if is_count_based_item(product, text):
        return "Net. 1 PCS"
    return None


def material_from_text(text: str) -> str | None:
    if not text:
        return None
    patterns = [
        r"(?:material|materials|made\s+of|made\s+from)\s*[:：]?\s*([A-Za-z0-9, /+\-().]{3,140})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if match:
            raw = re.split(
                r"\b(?:size|weight|dimensions?|package|country|made in|warning|caution|description|shipping)\b",
                match.group(1),
                maxsplit=1,
                flags=re.I,
            )[0]
            material = normalize_material_text(raw)
            if material:
                return material
    material_terms = [
        ("ATBC-PVC", r"\bATBC[-\s]?PVC\b"),
        ("PVC", r"\bPVC\b"),
        ("ABS", r"\bABS\b"),
        ("PBT", r"\bPBT\b"),
        ("Nylon", r"\bNylon\b"),
        ("Silicone", r"\bSilicone\b"),
        ("Polyester", r"\bPolyester\b"),
        ("Stainless Steel", r"\bStainless\s+Steel\b"),
        ("Plastic", r"\bPlastic\b"),
    ]
    for label, pattern in material_terms:
        if re.search(pattern, text, flags=re.I):
            return label
    return None


def normalize_material_text(material: str) -> str | None:
    material = html.unescape(str(material or ""))
    material = re.sub(r"<[^>]+>", " ", material)
    material = re.sub(r"\s+", " ", material).strip(" .;:-")
    if not material or len(material) < 3:
        return None
    low = material.lower()
    if "atbc" in low and "pvc" in low:
        return "ATBC-PVC"
    replacements = {
        "pvc": "PVC",
        "abs": "ABS",
        "pbt": "PBT",
    }
    parts = [part.strip(" .;:-") for part in re.split(r",|/|\+", material) if part.strip(" .;:-")]
    cleaned = []
    for part in parts[:6]:
        normalized = replacements.get(part.lower(), part)
        cleaned.append(normalized)
    return ", ".join(dict.fromkeys(cleaned)) if cleaned else None


def material_label_from_text(product: str, text: str = "") -> str | None:
    if not is_count_based_item(product, text):
        return None
    material = material_from_text(text)
    if not material:
        return None
    return f"MATERIAL/MATÉRIAU: {material}"


def is_material_label(value: str) -> bool:
    clean = searchable_text(value)
    return clean.startswith("material materiau") or clean.startswith("material")


def coo_from_text(text: str) -> str | None:
    country_map = {
        "korea": "Made In Korea / Fabriqué En Corée",
        "south korea": "Made In Korea / Fabriqué En Corée",
        "japan": "Made In Japan / Fabriqué au Japon",
        "china": "Made In China / Fabriqué En Chine",
    }
    low = text.lower()
    for key, value in country_map.items():
        if (
            f"made in {key}" in low
            or f"product of {key}" in low
            or f"country of origin {key}" in low
            or f"country/region of origin {key}" in low
        ):
            return value
    if "chinese makeup brand" in low or "china makeup" in low or "chinese cosmetics" in low:
        return country_map["china"]
    return None


def domain_from_url(url: str) -> str:
    domain = urlparse(url).netloc.lower()
    return domain.removeprefix("www.")


def ingredients_from_text(text: str, product: str = "") -> str | None:
    if not text:
        return None
    text = html.unescape(html.unescape(text))
    text = text.replace("\\n", " ").replace('\\"', '"')
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    shade_ingredients = ingredients_for_shade(text, product)
    if shade_ingredients:
        label = ingredients_label(shade_ingredients)
        return None if label == "need to review" else label
    stop_words = (
        r"\bmore\b|more information|this list of ingredients|actual ingredients|"
        r"ingredients subject|shipping policy|policies|"
        r"ingredient-list-copy|copy find dupes|find dupes|discover better matches|"
        r"key ingredients|ingredients explained|benefits|"
        r"product information|product details|details\s*/|"
        rf"{USAGE_HEADING_RE}|mode d’emploi|{CAUTION_HEADING_RE}|made in|catalog|sku|size|{NET_WEIGHT_HEADING_RE}"
    )
    patterns = [
        rf"{INGREDIENT_HEADING_RE}\s*[:：]?\s*(.{{40,9000}}?)(?=\s*(?:{stop_words})|$)",
    ]
    candidates: list[str] = []
    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.I):
            candidate = normalize_ingredients_text(match.group(1))
            if len(candidate) >= 40 and ("," in candidate or looks_like_uncomma_ingredient_list(candidate)):
                candidates.append(candidate)
    candidates = [
        candidate
        for candidate in candidates
        if valid_ingredient_candidate(candidate)
    ]
    if not candidates:
        return None
    ingredients = max(candidates, key=ingredient_candidate_score)
    if ingredient_candidate_score(ingredients) < 4:
        return None
    label = ingredients_label(ingredients)
    return None if label == "need to review" else label


def ingredients_for_shade(text: str, product: str) -> str | None:
    shade_pattern = shade_regex(product)
    if not shade_pattern:
        return None
    ingredient_head = INGREDIENT_HEADING_RE
    next_shade_pattern = r"(?<![-/\w])#?\s*(?:\d{2,3}|[a-z]{1,2}\s*\d{1,3})\s+(?-i:[A-Z][A-Za-z]+)"
    candidates = []
    for head_match in re.finditer(ingredient_head, text, flags=re.I):
        section = text[head_match.end(): head_match.end() + 9000]
        for shade_match in re.finditer(shade_pattern, section, flags=re.I):
            tail = section[shade_match.end():]
            next_match = re.search(next_shade_pattern, tail, flags=re.I)
            candidate_text = tail[: next_match.start()] if next_match else tail
            candidate = normalize_ingredients_text(candidate_text)
            if (
                len(candidate) >= 40
                and ("," in candidate or looks_like_uncomma_ingredient_list(candidate))
                and valid_ingredient_candidate(candidate)
                and ingredient_candidate_score(candidate) >= 4
            ):
                candidates.append(candidate)
    if candidates:
        return max(candidates, key=ingredient_candidate_score)

    start_pattern = rf"{ingredient_head}.*?{shade_pattern}\s*[:：]?\s*"
    stop_pattern = (
        r"(?=\s*(?:(?<![-/\w])#?\s*(?:\d{2,3}|[a-z]{1,2}\s*\d{1,3})\s+(?-i:[A-Z][A-Za-z]+)|"
        r"\bmore\b|this list of ingredients|actual ingredients|ingredients subject|"
        r"ingredient-list-copy|copy find dupes|find dupes|discover better matches|"
        r"key ingredients|ingredients explained|benefits|"
        r"product information|shipping policy|"
        rf"{USAGE_HEADING_RE}|{CAUTION_HEADING_RE}|{NET_WEIGHT_HEADING_RE}|made in)|$)"
    )
    for match in re.finditer(start_pattern + rf"(.{{40,3000}}?){stop_pattern}", text, flags=re.I):
        candidate = normalize_ingredients_text(match.group(1))
        if (
            len(candidate) >= 40
            and ("," in candidate or looks_like_uncomma_ingredient_list(candidate))
            and valid_ingredient_candidate(candidate)
            and ingredient_candidate_score(candidate) >= 4
        ):
            candidates.append(candidate)
    if candidates:
        return max(candidates, key=ingredient_candidate_score)
    return None


def ingredient_candidate_score(ingredients: str) -> int:
    low = ingredients.lower()
    inci_hits = [
        "ci ",
        "dimethicone",
        "isostearate",
        "talc",
        "mica",
        "silica",
        "wax",
        "oil",
        "aqua",
        "glyceryl",
        "polyglyceryl",
        "cyclopentasiloxane",
        "hexanediol",
    ]
    return sum(low.count(hit) for hit in inci_hits) + ingredients.count(",")


def valid_ingredient_candidate(ingredients: str) -> bool:
    low = ingredients.lower()
    bad_markers = [
        "manufacturer's discretion",
        "refer to product packaging",
        "subject to change",
        "productdata.",
        "customer reviews",
        "how to use",
        "directions",
        "direction for use",
        "apply ",
        "use a brush",
        "using a brush",
        "fingertip",
        "blend evenly",
        "product details",
        "product information",
        "official service",
        "login register",
        "decode inci",
        "products ingredients decode",
        "follow us on",
        "we don't have description",
        "what-it-does",
        "also-called",
        "viscosity controlling",
        "helper ingredient",
        "read what notable effects",
        "shipping",
        "return policy",
        "add to cart",
        "suitable for",
        "recommended for",
        "benefits",
    ]
    if "{{" in ingredients or any(marker in low for marker in bad_markers):
        return False
    if ingredients.count(",") < 2 and not looks_like_uncomma_ingredient_list(ingredients):
        return False
    if ingredient_candidate_score(ingredients) < 4:
        return False
    parts = split_ingredients(ingredients)
    if len(parts) < 4 and not looks_like_uncomma_ingredient_list(ingredients):
        return False
    prose_words = [
        "apply",
        "blend",
        "skin",
        "lip",
        "cheek",
        "brush",
        "fingertips",
        "texture",
        "finish",
        "color",
        "shade",
    ]
    prose_hits = sum(1 for word in prose_words if re.search(rf"\b{re.escape(word)}\b", low))
    if looks_like_uncomma_ingredient_list(ingredients):
        return prose_hits <= 2
    inci_like = sum(
        1
        for part in parts
        if re.search(
            r"\b(?:ci\s*\d{5}|aqua|dimethicone|mica|talc|silica|wax|oil|extract|glycol|"
            r"glycerin|glycerol|isostearate|stearate|methicone|poly|copolymer|oxide|"
            r"titanium|iron|phenoxyethanol|ethylhexylglycerin|tocopherol)\b",
            part,
            flags=re.I,
        )
    )
    return inci_like >= 2 and prose_hits <= max(2, len(parts) // 8)


def looks_like_uncomma_ingredient_list(ingredients: str) -> bool:
    low = ingredients.lower()
    words = re.findall(r"[a-z][a-z0-9/-]+", low)
    if len(words) < 18:
        return False
    markers = [
        "silica",
        "isononyl",
        "polybutene",
        "ethylhexyl",
        "synthetic wax",
        "propylene glycol",
        "triethylhexanoin",
        "diisostearyl",
        "microcrystalline wax",
        "titanium dioxide",
        "polyglyceryl",
        "candelilla",
        "dimethicone",
        "octyldodecanol",
        "iron oxides",
        "tocopheryl",
        "ci ",
    ]
    hits = sum(1 for marker in markers if marker in low)
    return hits >= 5 and bool(re.search(r"\bCI\s*\d{5}\b|\(CI\s*\d{5}\)", ingredients, flags=re.I))


def normalize_ingredients_text(ingredients: str) -> str:
    ingredients = re.sub(r'^[\w\s"{}:,-]*["\']ingredients["\']\s*:\s*["\']?', "", ingredients, flags=re.I)
    ingredients = re.sub(r"^(?:major\s+)?ingredients?\s*(?:[:：-])?\s*", "", ingredients, flags=re.I)
    ingredients = re.sub(r"^ingr[eé]dients?\s*(?:[:：-])?\s*", "", ingredients, flags=re.I)
    ingredients = trim_non_ingredient_tail(ingredients)
    ingredients = re.sub(
        r"^(?:[A-Z][A-Z0-9'/-]*\s+){1,4}(?=(?:Aqua|Water|Silica|Dimethicone|Mica|Talc|"
        r"Isononyl|Polybutene|Ethylhexyl|Synthetic|Titanium|Iron|CI)\b)",
        "",
        ingredients,
    )
    ingredients = re.sub(r"\bFormula\s+\d+\s*:\s*", ", ", ingredients, flags=re.I)
    ingredients = re.sub(r"\b(GL|PG|VT)\d{1,3}\b\s*", "", ingredients)
    ingredients = re.sub(r"(?<![A-Za-z])(?:C|W|N)\d{1,3}(?![-\w])\s*", "", ingredients)
    ingredients = re.sub(r"\s*<br\s*/?>\s*", ", ", ingredients, flags=re.I)
    ingredients = re.sub(r"\s*,\s*", ", ", ingredients)
    ingredients = re.sub(r"\s+", " ", ingredients).strip(" .;")
    ingredients = re.sub(r"\bwater\b", "Aqua", ingredients, flags=re.I)
    ingredients = re.sub(r"\bINCI\b\s*$", "", ingredients).strip(" .;,")
    ingredients = dedupe_adjacent_ingredients(ingredients)
    return ingredients


def trim_non_ingredient_tail(ingredients: str) -> str:
    markers = [
        "please note:",
        "please note",
        "note:",
        "ingredients may differ",
        "may differ depending on shade",
        "may vary by shade",
        "how to use",
        "directions",
        "direction for use",
        "product details",
        "product information",
        "description:",
        "product description",
        "official service",
        "ingredient-list-copy",
        "copy find dupes",
        "copy",
        "find dupes",
        "discover better matches",
        "key ingredients",
        "ingredients explained",
        "show all",
        "show less",
        "read more",
        "customer reviews",
        "review",
        "rating",
        "add to cart",
        "add to bag",
        "wishlist",
        "select shade",
        "select color",
        "shipping",
        "return policy",
        "login register",
        "decode inci",
        "products ingredients decode",
        "follow us on",
        "skin conditioning, solvent",
        "supports skin hydration",
        "shields skin",
        "benefits hydrating",
        "class=\"inline-flex",
    ]
    low = ingredients.lower()
    cut_at = len(ingredients)
    for marker in markers:
        idx = low.find(marker)
        if idx != -1:
            cut_at = min(cut_at, idx)
    return ingredients[:cut_at].strip(" .;,/>")


def dedupe_adjacent_ingredients(ingredients: str) -> str:
    parts = split_ingredients(ingredients)
    cleaned: list[str] = []
    seen: set[str] = set()
    for part in parts:
        key = ingredient_dedupe_key(part)
        if key and key not in seen:
            cleaned.append(part.strip())
            seen.add(key)
    return ", ".join(cleaned)


def ingredient_dedupe_key(ingredient: str) -> str:
    key = re.sub(r"\s+", " ", ingredient).strip().lower()
    key = key.replace("ci ", "ci")
    key = re.sub(r"[^a-z0-9]+", "", key)
    return key


def strip_ingredients_label_prefix(ingredients: str) -> str:
    text = str(ingredients or "").strip()
    text = re.sub(
        r"^INGREDIENTS/(?:INGR[ÉE]?.?DIENTS|INGREDIENTS IN FRENCH)\s*:\s*",
        "",
        text,
        flags=re.I,
    )
    text = re.sub(r"^INGREDIENTS\s*:\s*", "", text, flags=re.I)
    return re.split(r"\s+/\s+", text, maxsplit=1)[0].strip()


def english_ingredients_text(ingredients: str | None) -> str:
    clean = extract_clean_ingredient_segment(strip_ingredients_label_prefix(str(ingredients or "")))
    if not clean:
        return ""
    if looks_french_ingredients(clean):
        clean = translate_ingredients(clean, FR_TO_EN_INGREDIENTS)
    return extract_clean_ingredient_segment(clean)


def extract_clean_ingredient_segment(ingredients: str | None) -> str:
    raw = str(ingredients or "").strip()
    clean = normalize_ingredients_text(raw)
    if not clean:
        clean = ""
    if valid_ingredient_output(clean):
        return clean

    candidates = [clean]
    first_ingredient = re.search(
        r"\b(?:Aqua|Water|Dimethicone|Silica|Mica|Talc|Glycerin|Butylene Glycol|"
        r"Polyglyceryl|Isododecane|Cyclopentasiloxane|Titanium Dioxide|Iron Oxides|CI\s*\d{5})\b",
        raw,
        flags=re.I,
    )
    if first_ingredient:
        candidates.append(raw[first_ingredient.start():])

    parts = split_ingredients(clean)
    if len(parts) >= 4:
        rebuilt: list[str] = []
        for part in parts:
            part_clean = trim_non_ingredient_tail(part).strip(" .;,")
            if not part_clean:
                break
            if ingredient_part_looks_dirty(part_clean):
                break
            rebuilt.append(part_clean)
        if len(rebuilt) >= 4:
            candidates.append(", ".join(rebuilt))

    for candidate in candidates[1:]:
        normalized = normalize_ingredients_text(candidate)
        if normalized and valid_ingredient_output(normalized):
            return normalized
    return ""


def ingredient_part_looks_dirty(part: str) -> bool:
    low = part.lower()
    dirty_markers = [
        "how to use",
        "direction",
        "product information",
        "product details",
        "description:",
        "customer review",
        "add to cart",
        "shipping",
        "return policy",
        "login",
        "register",
        "decode inci",
        "copy",
        "find dupes",
        "ingredients explained",
        "we don't have description",
        "what-it-does",
        "also-called",
        "helper ingredient",
        "http://",
        "https://",
        ".com",
    ]
    return any(marker in low for marker in dirty_markers)


def clean_ingredients_for_output(ingredients: str | None) -> str:
    clean = english_ingredients_text(ingredients)
    if not clean:
        return ""
    if not valid_ingredient_output(clean):
        return ""
    return clean


def valid_ingredient_output(ingredients: str) -> bool:
    low = ingredients.lower()
    bad_markers = [
        "major ingredients",
        "ingredients:",
        "ingrédients:",
        "how to use",
        "direction for use",
        "product information",
        "product details",
        "description:",
        "customer review",
        "add to cart",
        "add to bag",
        "wishlist",
        "select shade",
        "select color",
        "shipping",
        "return policy",
        "login",
        "register",
        "decode inci",
        "copy",
        "find dupes",
        "key ingredients",
        "ingredients explained",
        "we don't have description",
        "what-it-does",
        "also-called",
        "viscosity controlling",
        "helper ingredient",
        "read what notable effects",
        "benefits",
        "http://",
        "https://",
        ".com",
    ]
    if any(marker in low for marker in bad_markers):
        return False
    if looks_like_uncomma_ingredient_list(ingredients):
        return valid_ingredient_candidate(ingredients)
    parts = split_ingredients(ingredients)
    if len(parts) < 4:
        return False
    if any(len(part) > 140 for part in parts):
        return False
    sentence_markers = [" apply ", " use ", " helps ", " leaves ", " provides ", " designed ", " suitable "]
    if sum(1 for marker in sentence_markers if marker in f" {low} ") >= 2:
        return False
    return valid_ingredient_candidate(ingredients)


def ingredients_label(ingredients: str | None) -> str:
    if not ingredients:
        return "need to review"
    english = clean_ingredients_for_output(ingredients)
    if not english:
        return "need to review"
    return f"{INGREDIENTS_OUTPUT_PREFIX} {english}"


def ingredients_label_from_known(known: dict[str, str]) -> str:
    if known.get("ingredients"):
        return ingredients_label(known.get("ingredients"))
    if known.get("ingredients_fr"):
        return ingredients_label(known.get("ingredients_fr"))
    return "need to review"


def bilingual_ingredients(ingredients: str) -> tuple[str, str]:
    clean = normalize_ingredients_text(ingredients)
    if looks_french_ingredients(clean):
        french = clean
        english = translate_ingredients(clean, FR_TO_EN_INGREDIENTS)
    else:
        english = clean
        french = translate_ingredients(clean, EN_TO_FR_INGREDIENTS)
    return english, french


def looks_french_ingredients(text: str) -> bool:
    low = text.lower()
    markers = ["diméthicone", "polymère", "croisé", "triisostéarate", "extrait", "huile", "acétate"]
    return any(marker in low for marker in markers)


EN_TO_FR_INGREDIENTS = {
    "Isostearyl Isostearate": "Isostéarate d’isostéaryle",
    "Polyglyceryl-2 Triisostearate": "Triisostéarate de polyglycéryle-2",
    "Diisostearyl Malate": "Malate de diisostéaryle",
    "Sorbitan Isostearate": "Isostéarate de sorbitan",
    "Paraffin": "Paraffine",
    "Trimethylpentaphenyl Trisiloxane": "Triméthylpentaphényl trisiloxane",
    "Microcrystalline Wax": "Cire microcristalline",
    "Pentaerythrityl Isostearate": "Isostéarate de pentaérythrityle",
    "Euphorbia Cerifera (Candelilla) Wax": "Cire d’Euphorbia cerifera (candelilla)",
    "1,2-Hexanediol": "1,2-hexanediol",
    "PEG/PPG-10/1 Dimethicone": "PEG/PPG-10/1 diméthicone",
    "Fragrance": "Parfum",
    "Pentaerythrityl Tetraisostearate": "Tétraisostéarate de pentaérythrityle",
    "Dimethicone": "Diméthicone",
    "Dimethicone Crosspolymer": "Polymère croisé de diméthicone",
    "Polydimethylsiloxane": "Polydiméthylsiloxane",
    "Cyclopentasiloxane": "Cyclopentasiloxane",
    "Cyclohexasiloxane": "Cyclohexasiloxane",
    "Silica Silylate": "Silylate de silice",
    "Cetyl PEG/PPG-10/1 Dimethicone": "Cétyl PEG/PPG-10/1 diméthicone",
    "Talc": "Talc",
    "Melaleuca Alternifolia Extract": "Extrait de Melaleuca alternifolia",
    "Evening Primrose Oil": "Huile d’onagre",
    "Alpha-Tocopheryl Acetate": "Acétate d’alpha-tocophéryle",
    "C30-45 Alkyl Dimethicone": "C30-45 Alkyl Diméthicone",
    "Talc": "Talc",
    "Mica": "Mica",
    "Synthetic Fluorphlogopite": "Fluorphlogopite synthétique",
    "Silica": "Silice",
    "Magnesium Myristate": "Myristate de magnésium",
    "Vinyl Dimethicone/Methicone Silsesquioxane Crosspolymer": "Polymère croisé vinyl diméthicone/méthicone silsesquioxane",
    "Octyldodecanol": "Octyldodécanol",
    "Isostearyl Neopentanoate": "Néopentanoate d’isostéaryle",
    "Methyl Methacrylate Crosspolymer": "Polymère croisé de méthacrylate de méthyle",
    "Aluminum Myristate": "Myristate d’aluminium",
    "Ethylhexylglycerin": "Éthylhexylglycérine",
    "Glyceryl Caprylate": "Caprylate de glycéryle",
    "Lauroyl Lysine": "Lauroyl lysine",
    "Triethoxycaprylylsilane": "Triéthoxycaprylylsilane",
    "Cocos Nucifera (Coconut) Oil": "Huile de Cocos nucifera (noix de coco)",
    "Alumina": "Alumine",
    "Aluminum Hydroxide": "Hydroxyde d’aluminium",
    "Boron Nitride": "Nitrure de bore",
    "Aqua": "Aqua",
    "Trimethylsiloxysilicate": "Trimethylsiloxysilicate",
    "Methyl Trimethicone": "Methyl trimethicone",
    "Candelilla Wax Hydrocarbon": "Hydrocarbure de cire de candelilla",
    "Hydrogenated Polyisobutene": "Polyisobutene hydrogene",
    "Synthetic Wax": "Cire synthetique",
    "Acrylates/Stearyl Acrylate/Dimethicone Methacrylate Copolymer": "Copolymere acrylates/acrylate de stearyle/methacrylate de dimethicone",
    "Behenyl Alcohol": "Alcool behenylique",
    "Tri(Behenic Acid/Isostearic Acid/Eicosanedioic Acid) Glyceryl": "Tri(behenate/isostearate/eicosanedioate) de glyceryle",
    "Ricinus Communis (Castor) Seed Oil": "Huile de graines de Ricinus communis (ricin)",
    "Squalane": "Squalane",
    "Butyrospermum Parkii (Shea) Butter": "Beurre de Butyrospermum parkii (karite)",
    "Rosa Canina Fruit Oil": "Huile de fruit de Rosa canina",
    "Persea Gratissima (Avocado) Oil": "Huile de Persea gratissima (avocat)",
    "Camellia Japonica Seed Oil": "Huile de graines de Camellia japonica",
    "Argania Spinosa Kernel Oil": "Huile de noyau d'Argania spinosa",
    "Simmondsia Chinensis (Jojoba) Seed Oil": "Huile de graines de Simmondsia chinensis (jojoba)",
    "Macadamia Ternifolia Seed Oil": "Huile de graines de Macadamia ternifolia",
    "Prinsepia Utilis Seed Oil": "Huile de graines de Prinsepia utilis",
    "Limnanthes Alba (Meadowfoam) Seed Oil": "Huile de graines de Limnanthes alba",
    "Carthamus Tinctorius (Safflower) Seed Oil": "Huile de graines de Carthamus tinctorius",
    "Helianthus Annuus (Sunflower) Seed Oil": "Huile de graines d'Helianthus annuus",
    "Rosa Damascena Flower Extract": "Extrait de fleur de Rosa damascena",
    "Lavandula Angustifolia (Lavender) Flower Extract": "Extrait de fleur de Lavandula angustifolia",
    "Tocopherol": "Tocopherol",
    "Tocopheryl Acetate": "Acetate de tocopheryle",
    "Iron Oxides": "Oxydes de fer",
    "Titanium Dioxide": "Dioxyde de titane",
}

FR_TO_EN_INGREDIENTS = {
    "Diméthicone": "Dimethicone",
    "polymère croisé de diméthicone": "Dimethicone Crosspolymer",
    "triisostéarate de polyglycéryle-2": "Polyglyceryl-2 Triisostearate",
    "cyclopentasiloxane": "Cyclopentasiloxane",
    "cyclohexasiloxane": "Cyclohexasiloxane",
    "1,2-hexanediol": "1,2-Hexanediol",
    "silylate de silice": "Silica Silylate",
    "cétyl PEG/PPG-10/1 diméthicone": "Cetyl PEG/PPG-10/1 Dimethicone",
    "talc": "Talc",
    "extrait de Melaleuca Alternifolia": "Melaleuca Alternifolia Extract",
    "huile d'onagre": "Evening Primrose Oil",
    "huile d’onagre": "Evening Primrose Oil",
    "acétate d'alpha-tocophéryle": "Alpha-Tocopheryl Acetate",
    "acétate d’alpha-tocophéryle": "Alpha-Tocopheryl Acetate",
    "C30-45 Alkyl Diméthicone": "C30-45 Alkyl Dimethicone",
}


def translate_ingredients(ingredients: str, mapping: dict[str, str]) -> str:
    parts = split_ingredients(ingredients)
    translated = []
    lower_map = {key.lower(): value for key, value in mapping.items()}
    for part in parts:
        translated.append(lower_map.get(part.lower(), mapping.get(part, part)))
    return ", ".join(translated)


def split_ingredients(ingredients: str) -> list[str]:
    protected = re.sub(r"(?<!\d)(\d),\s*(\d)(?=[-\w])", r"\1<COMMA>\2", ingredients)
    parts = [part.replace("<COMMA>", ",").strip() for part in protected.split(",")]
    return [part for part in parts if part]


def manufacturer_from_text(text: str) -> str | None:
    patterns = [
        r"manufacturer\s*[:：]\s*(.{3,120}?)(?:address|country|made in|$)",
        r"company\s+name\s*[:：]\s*(.{3,120}?)(?:country|address|business|$)",
        r"name\s+of\s+business/corporation\s*[:：]\s*(.{3,120}?)(?:address|$)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if match:
            value = re.sub(r"\s+", " ", match.group(1)).strip(" .;")
            if len(value) > 3:
                return value
    return None


def how_to_use_from_text(text: str) -> str | None:
    match = re.search(
        rf"{USAGE_HEADING_RE}\s*[:：]?\s*(.{{20,700}}?)(?:{INGREDIENT_HEADING_RE}|{NET_WEIGHT_HEADING_RE}|official service|shipping|{CAUTION_HEADING_RE}|customer|$)",
        text,
        flags=re.I,
    )
    if not match:
        return None
    raw = re.sub(r"\s+", " ", match.group(1)).strip(" .:")
    if len(raw) < 15:
        return None
    return raw


def product_name_fr(product: str) -> str:
    clean = expand_product_abbreviations(normalize_product_name(product))
    clean = re.sub(r"^\d{5,14}\s+", "", clean)
    if is_love_liner_liquid_eyeliner(clean):
        color = ""
        if love_liner_color(clean):
            color = " " + SHADE_FR_TERMS.get(love_liner_color(clean), love_liner_color(clean))
        return f"Love Liner traceur liquide pour les yeux R5{color}".strip()
    if is_love_liner_cream_fit(clean):
        variant = " ovale ultra fin" if love_liner_variant(clean) == "Ultra Slim" else ""
        color = ""
        if love_liner_color(clean):
            color = " " + SHADE_FR_TERMS.get(love_liner_color(clean), love_liner_color(clean))
        return f"Love Liner crayon pour les yeux Cream Fit R{variant}{color}".strip()
    result = translate_product_terms(clean)
    if "smiski" in searchable_text(clean):
        result = re.sub(r"\bbain\b", "série bain", result, flags=re.I)
    result = re.sub(r"(\d+(?:\.\d+)?)\s*g\b", r"\1 g", result, flags=re.I)
    result = re.sub(r"(\d+(?:\.\d+)?)\s*ml\b", r"\1 mL", result, flags=re.I)
    result = re.sub(r"\s*\+\s*", " + ", result)
    result = re.sub(r"\)(?=\w)", ") ", result)
    return re.sub(r"\s+", " ", result).strip()


def translate_product_terms(product: str) -> str:
    result = product
    for mapping in (PRODUCT_NAME_FR_PHRASES, SHADE_FR_TERMS):
        for english, french in sorted(mapping.items(), key=lambda item: len(item[0]), reverse=True):
            result = replace_phrase_case_insensitive(result, english, french)
    return result


def replace_phrase_case_insensitive(text: str, english: str, french: str) -> str:
    escaped = re.escape(english).replace(r"\ ", r"\s+")
    pattern = rf"(?<![A-Za-z]){escaped}(?![A-Za-z])"
    return re.sub(pattern, french, text, flags=re.I)


def french_name_needs_translation(value: str) -> bool:
    clean = searchable_text(value)
    english_terms = [
        "lip",
        "cheek",
        "glowy",
        "jelly",
        "pocket",
        "eye palette",
        "eyeshadow",
        "lip tint",
        "voluming",
        "gloss",
        "highlighter",
        "pencil",
        "eyeliner",
        "figure",
        "figures",
        "blind box",
        "mystery box",
        "random style",
        "brush",
        "applicator",
        "sponge",
        "puff",
        "glow",
        "bath",
    ]
    return any(term in clean for term in english_terms)


def normalize_product_name(product: str) -> str:
    return re.sub(r"\s+", " ", str(product).replace("\xa0", " ")).strip()


def compact_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_product_name(value).lower())


def searchable_text(value: str) -> str:
    text = normalize_product_name(value).lower()
    text = re.sub(r"[\-_()/|]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def contains_search_term(value: str, term: str) -> bool:
    clean = searchable_text(value)
    escaped = re.escape(term.lower()).replace(r"\ ", r"\s+")
    return bool(re.search(rf"(?<![a-z0-9]){escaped}(?![a-z0-9])", clean))


def contains_any_search_term(value: str, terms: list[str]) -> bool:
    return any(contains_search_term(value, term) for term in terms)


def is_count_based_item(product: str, source_text: str = "") -> bool:
    if contains_any_search_term(product, COUNT_ITEM_TERMS):
        return True
    return bool(source_text and contains_any_search_term(source_text, COLLECTIBLE_TERMS))


def is_collectible_item(product: str, source_text: str = "") -> bool:
    if contains_any_search_term(product, COLLECTIBLE_TERMS):
        return True
    return bool(source_text and contains_any_search_term(source_text, COLLECTIBLE_TERMS))


def is_tool_item(product: str, source_text: str = "") -> bool:
    return contains_any_search_term(product, TOOL_ITEM_TERMS)


def expand_product_abbreviations(product: str) -> str:
    result = normalize_product_name(product)
    compact = compact_text(result)
    for phrase, replacement in sorted(PRODUCT_PHRASE_ALIASES.items(), key=lambda item: len(item[0]), reverse=True):
        if " " in phrase or "&" in phrase:
            result = replace_phrase_case_insensitive(result, phrase, replacement)
        elif phrase in compact:
            result = re.sub(phrase, replacement, result, flags=re.I)
    for abbr, full in PRODUCT_ABBREVIATIONS.items():
        prefix_guard = r"(?<![A-Za-z0-9&])" if abbr == "BE" else r"(?<![A-Za-z0-9])"
        result = re.sub(rf"{prefix_guard}{re.escape(abbr)}(?![A-Za-z0-9])", full, result, flags=re.I)
    return re.sub(r"\s+", " ", result).strip()


def brand_rule_for_product(product: str) -> dict[str, Any] | None:
    clean = searchable_text(expand_product_abbreviations(product))
    compact = compact_text(product)
    for rule in BRAND_ALIAS_RULES:
        if all(keyword in clean or keyword in compact for keyword in rule["keywords"]):
            return rule
        if any(re.search(pattern, clean, flags=re.I) or re.search(pattern, compact, flags=re.I) for pattern in rule["patterns"]):
            return rule
    return None


def remove_brand_alias(product: str, rule: dict[str, Any] | None) -> str:
    result = normalize_product_name(product)
    if not rule:
        return result
    for pattern in rule["patterns"]:
        result = re.sub(pattern, " ", result, count=1, flags=re.I)
    return re.sub(r"\s+", " ", result).strip(" -_")


def brand_domains_for_product(product: str) -> list[str]:
    rule = brand_rule_for_product(product)
    if not rule:
        return []
    return list(rule.get("domains", []))


def generic_product_alias_names(product: str) -> list[str]:
    base = normalize_product_name(product)
    expanded = expand_product_abbreviations(base)
    names = [base, expanded]
    rule = brand_rule_for_product(expanded)
    if rule:
        tail = remove_brand_alias(expanded, rule)
        for alias in rule.get("aliases", []):
            names.append(f"{alias} {tail}".strip())
    return names


def generic_color_name(product: str) -> str:
    expanded = expand_product_abbreviations(product)
    clean = searchable_text(expanded)
    for color in [
        "Medium Brown",
        "Dark Brown",
        "Ash Brown",
        "Black",
        "Brown",
        "Red",
        "Pink",
        "Coral",
        "Orange",
        "Beige",
        "Gray",
        "White",
    ]:
        if searchable_text(color) in clean:
            return color
    return ""


def is_love_liner_cream_fit(product: str) -> bool:
    clean = searchable_text(expand_product_abbreviations(product))
    compact = compact_text(product)
    brand_match = "loveliner" in compact or "loveerliner" in compact or "love liner" in clean
    cream_fit_match = "creamfit" in compact or ("cream" in clean and "fit" in clean)
    return brand_match and cream_fit_match and ("pencil" in clean or "liner" in clean)


def is_love_liner_liquid_eyeliner(product: str) -> bool:
    clean = searchable_text(expand_product_abbreviations(product))
    compact = compact_text(product)
    brand_match = "loveliner" in compact or "loveerliner" in compact or "love liner" in clean
    liquid_match = "liquideyeliner" in compact or ("liquid" in clean and ("eyeliner" in clean or "eye liner" in clean))
    return brand_match and liquid_match


def love_liner_color(product: str) -> str:
    clean = searchable_text(product)
    compact = compact_text(product)
    if "mediumbrown" in compact or re.search(r"\bmbr\b", clean):
        return "Medium Brown"
    if "ashbrown" in compact or re.search(r"\babr\b", clean):
        return "Ash Brown"
    if "black" in clean or re.search(r"\bbk\b", clean):
        return "Black"
    return generic_color_name(product)


def love_liner_variant(product: str) -> str:
    clean = searchable_text(product)
    compact = compact_text(product)
    if "slimoval" in compact or "ultra slim" in clean or "slim oval" in clean:
        return "Ultra Slim"
    return ""


def expanded_product_names(product: str) -> list[str]:
    base = normalize_product_name(product)
    names = generic_product_alias_names(base)
    if is_love_liner_cream_fit(base):
        color = love_liner_color(base)
        variant = love_liner_variant(base)
        color_tail = f" {color}" if color else ""
        variant_tail = f" {variant}" if variant else ""
        names.extend(
            [
                f"MSH Love Liner Cream Fit Pencil R{variant_tail}{color_tail}",
                f"LOVE Liner Cream Fit Pencil Liner{variant_tail}{color_tail}",
                f"Love Liner Cream Fit Pencil R Slimoval{color_tail}",
                f"Love Liner Cream Fit Pencil R SlimOval MBR",
                f"Love Liner Cream Fit Pencil R Slimoval Medium Brown 0.05g",
                f"MSH Love Liner Cream Fit Pencil{color_tail}",
            ]
        )
    if is_love_liner_liquid_eyeliner(base):
        color = love_liner_color(base) or "Black"
        color_tail = f" {color}" if color else ""
        names.extend(
            [
                f"MSH Love Liner Liquid Eyeliner R5{color_tail}",
                f"Love Liner Liquid Eyeliner R5{color_tail} 0.55ml",
                f"Love Liner Liquid Eyeliner R5 {color}",
                f"msh Love Liner Liquid Eyeliner R5 {color}",
                f"4570159423723 Love Liner Liquid Eyeliner R5 {color}",
            ]
        )
    deduped: list[str] = []
    seen = set()
    for name in names:
        name = re.sub(r"\s+", " ", name).strip()
        key = name.lower()
        if name and key not in seen:
            seen.add(key)
            deduped.append(name)
    return deduped


def search_tokens(product: str) -> list[str]:
    stopwords = {"the", "and", "for", "with", "colors", "color", "set", "pcs", "pc"}
    tokens = re.findall(r"[\w\u3040-\u30ff\u3400-\u9fff\uac00-\ud7af]+", searchable_text(product))
    return [token for token in tokens if len(token) > 1 and token not in stopwords]


def shade_key(product: str) -> str:
    clean = searchable_text(product)
    match = re.search(r"(?:#\s*|\b)([a-z]{1,3}\s*\d{1,3})\b\s+([a-z][a-z0-9]+(?:\s+[a-z][a-z0-9]+){0,2})", clean)
    if match:
        code = re.sub(r"\s+", "", match.group(1))
        shade = re.sub(r"\s+", "-", match.group(2).strip())
        return f"{code}-{shade}"
    match = re.search(r"(?:#\s*|\b)(\d{1,3})\b\s+([a-z][a-z0-9]+(?:\s+[a-z][a-z0-9]+){0,2})", clean)
    if match:
        shade = re.sub(r"\s+", "-", match.group(2).strip())
        return f"{match.group(1)}-{shade}"
    return ""


def shade_code(product: str) -> str:
    match = re.search(r"#\s*([a-z]{0,3}\s*\d{1,3})\b", str(product or ""), flags=re.I)
    if match:
        return re.sub(r"\s+", "", match.group(1)).lower()
    clean = searchable_text(product)
    match = re.search(r"\b(?:shade|color|colour)\s*([a-z]{0,3}\s*\d{1,3})\b", clean)
    if match:
        return re.sub(r"\s+", "", match.group(1)).lower()
    return ""


def shade_regex(product: str) -> str:
    key = shade_key(product)
    if key:
        code, _, shade = key.partition("-")
    else:
        code = shade_code(product)
        shade = ""
    if not code:
        return ""
    code_pattern = r"\s*".join(re.escape(ch) for ch in code)
    if code.isdigit() and len(code) == 2:
        prefixed_code = r"\s*".join(re.escape(ch) for ch in f"1{code}")
        code_pattern = rf"(?:{code_pattern}|{prefixed_code})"
    shade_pattern = r"\s+".join(re.escape(part) for part in shade.split("-") if part)
    if not shade_pattern:
        return rf"#?\s*{code_pattern}\b"
    return rf"#?\s*{code_pattern}\s+{shade_pattern}"


def product_family_key(product: str) -> str:
    clean = searchable_text(product)
    shade = shade_key(product)
    if shade:
        code, _, shade_name = shade.partition("-")
        shade_words = r"\s+".join(re.escape(part) for part in shade_name.split("-") if part)
        if shade_words:
            clean = re.sub(rf"#?\s*{re.escape(code)}\s+{shade_words}\b\s*$", "", clean)
    clean = re.sub(r"\b[a-z]{1,3}\s*\d{1,3}\b\s*$", "", clean)
    clean = re.sub(r"\b#?\d{1,3}\b\s*$", "", clean)
    clean = re.sub(r"\b(?:shade|color|colour)\s*[a-z0-9]+\b\s*$", "", clean)
    return re.sub(r"\s+", " ", clean).strip()


def cacheable_family_key(product: str) -> str:
    family = product_family_key(product)
    if not family:
        return ""
    tokens = family.split()
    brand_tokens = {
        "into",
        "you",
        "ity",
        "judydoll",
        "millefee",
        "fwee",
        "romand",
        "peripera",
        "smiski",
        "dreams",
    }
    category_tokens = {
        "lip",
        "lipstick",
        "mud",
        "blush",
        "serum",
        "palette",
        "highlighter",
        "eye",
        "eyeshadow",
        "shadow",
        "cheek",
        "tint",
        "balm",
        "gloss",
        "mascara",
        "liner",
        "eyeliner",
        "figure",
        "figures",
        "figurine",
        "figurines",
        "blind",
        "box",
        "mystery",
        "random",
        "brush",
        "applicator",
        "sponge",
        "puff",
        "tool",
    }
    series_tokens = [token for token in tokens if token not in brand_tokens and token not in category_tokens]
    return family if series_tokens else ""


def product_brand(product: str) -> str:
    rule = brand_rule_for_product(product)
    if rule and rule.get("aliases"):
        return rule["aliases"][0]
    tokens = search_tokens(product)
    if not tokens:
        return ""
    if len(tokens) >= 2 and tokens[0] in {"into", "fwee"}:
        return " ".join(tokens[:2]) if tokens[0] == "into" else tokens[0]
    return " ".join(tokens[:2])


def fuzzy_queries(barcode: str, product: str, search_mode: str = DEFAULT_SEARCH_MODE) -> list[str]:
    clean_product = normalize_product_name(product)
    aliases = expanded_product_names(clean_product)
    search_basis = aliases[1] if len(aliases) > 1 else clean_product
    brand = product_brand(clean_product)
    brand_tail = remove_brand_alias(search_basis, brand_rule_for_product(search_basis))
    tokens = search_tokens(brand_tail or search_basis)
    core = " ".join(tokens[:6])
    tail = " ".join(tokens[:6])
    brand_domains = brand_domains_for_product(clean_product)
    if search_mode == "Deep":
        with CACHE_LOCK:
            remembered = load_search_cache().get("brands", {}).get(brand, []) if brand else []
        if isinstance(remembered, list):
            for domain in remembered:
                if domain not in brand_domains:
                    brand_domains.append(domain)
    domain_queries = [f"{core} site:{domain}" for domain in brand_domains[:5] if core]
    queries = [
        f"\"{clean_product}\"",
        f"\"{search_basis}\"",
        *domain_queries,
        f"\"{clean_product}\" ingredients",
        f"\"{search_basis}\" ingredients",
        f"\"{search_basis}\" \"major ingredients\"",
        f"\"{search_basis}\" INCI",
        f"{search_basis} YesStyle ingredients",
        f"{search_basis} Hwahae ingredients",
        f"{search_basis} AsianBeautyWholesale ingredients",
        f"\"{search_basis}\" net weight",
        f"{search_basis} ingredients net weight",
        f"{search_basis} how to use ingredients",
        f"{search_basis} official",
        f"{search_basis} YesStyle",
        f"{search_basis} OliveYoung",
        f"{search_basis} Kiseki",
        f"{brand} {tail} ingredients" if brand and tail else "",
        f"{core} site:yesstyle.com",
        f"{core} site:oliveyoung.com",
    ]
    if is_count_based_item(clean_product):
        queries.extend(
            [
                f"{search_basis} blind box 1 pack",
                f"{search_basis} random style",
                f"{search_basis} material",
                f"{search_basis} official",
            ]
        )
    for alias in aliases[1:]:
        queries.extend(
            [
                f"\"{alias}\"",
                f"\"{alias}\" ingredients",
                f"\"{alias}\" net weight",
                f"{alias} YesStyle",
                f"{alias} OliveYoung",
                f"{alias} official",
            ]
        )
        for domain in brand_domains[:5]:
            queries.append(f"{alias} site:{domain}")
    if search_mode == "Fast":
        queries = queries[:16]
    if barcode:
        queries[3:3] = [
            f"{barcode} {search_basis}",
            f"{barcode} ingredients net weight",
            f"\"{barcode}\"",
        ]
    if re.search(r"[\u3040-\u30ff\u3400-\u9fff\uac00-\ud7af]", clean_product):
        queries.extend(
            [
                f"{clean_product} ingredients English",
                f"{clean_product} 成分 容量",
                f"{clean_product} 全成分 内容量",
                f"{clean_product} 전성분 용량",
            ]
        )
    for hint in SEARCH_LANGUAGE_HINTS:
        if brand and hint:
            queries.append(f"{brand} {hint} {tail}".strip())
    deduped: list[str] = []
    seen = set()
    for query in queries:
        query = re.sub(r"\s+", " ", query).strip()
        if query and query not in seen:
            seen.add(query)
            deduped.append(query)
    return deduped[:48]


def fuzzy_source_score(url: str, title_or_text: str, barcode: str, product: str) -> float:
    if is_noise_url(url):
        return -10.0
    haystack = searchable_text(f"{url} {title_or_text}")
    alias_names = expanded_product_names(product)
    tokens = []
    for alias in alias_names:
        tokens.extend(search_tokens(alias))
    tokens = list(dict.fromkeys(tokens))
    score = 0.0
    if barcode and barcode in haystack:
        score += 8
    domain_rank = source_rank(url)
    if domain_rank < 100:
        score += 6 - min(domain_rank, 5) * 0.5
    score += product_detail_bonus(url)
    for token in tokens:
        if token in haystack:
            score += 1
    shade = shade_key(product)
    if shade and shade in haystack:
        score += 5
    elif shade:
        score -= 2
    if "ingredients" in haystack or "major ingredients" in haystack:
        score += 2
    if "net weight" in haystack or re.search(r"\b\d+(?:\.\d+)?\s*(g|ml)\b", haystack):
        score += 1.5
    if is_count_based_item(product) and (
        "blind box" in haystack or "random style" in haystack or "1 pack" in haystack or "figure" in haystack
    ):
        score += 2
    product_norms = [searchable_text(alias) for alias in alias_names if alias]
    if product_norms and haystack:
        score += max(difflib.SequenceMatcher(None, product_norm[:120], haystack[:240]).ratio() for product_norm in product_norms) * 4
    color = generic_color_name(product)
    if color and searchable_text(color) in haystack:
        score += 2
    return score


def direction_for_product(product: str, source_direction: str | None, source_text: str = "") -> tuple[str, str]:
    low = searchable_text(f"{product} {source_text}")
    if is_collectible_item(product, source_text):
        if "glow" in low or "dark" in low or "smiski" in low:
            en = (
                "DIRECTION FOR USE: Open the blind box and display the figure as desired. "
                "Expose to light to activate the glow-in-the-dark effect."
            )
            fr = (
                "MODE D’EMPLOI: Ouvrir la boîte surprise et exposer la figurine selon vos préférences. "
                "L’exposer à la lumière pour activer l’effet phosphorescent."
            )
        else:
            en = "DIRECTION FOR USE: Open the package and display the item as desired."
            fr = "MODE D’EMPLOI: Ouvrir l’emballage et exposer l’article selon vos préférences."
    elif is_tool_item(product, source_text):
        en = "DIRECTION FOR USE: Use the tool to apply, blend, or groom as needed. Clean regularly and allow to dry."
        fr = "MODE D’EMPLOI: Utiliser l’outil pour appliquer, estomper ou soigner au besoin. Nettoyer régulièrement et laisser sécher."
    elif "blush serum" in low or ("liquid" in low and "blush" in low):
        en = "DIRECTION FOR USE: Lightly dab after foundation and before setting powder. Tap and blend evenly across cheeks."
        fr = "MODE D’EMPLOI: Tapoter légèrement après le fond de teint et avant la poudre fixatrice. Estomper uniformément sur les joues."
    elif "highlighter" in low:
        en = "DIRECTION FOR USE: Apply to high points of the face with a brush. Layer as desired for added glow."
        fr = "MODE D’EMPLOI: Appliquer sur les points saillants du visage à l’aide d’un pinceau. Superposer au besoin pour plus d’éclat."
    elif "eyeliner" in low or ("eye" in low and "liner" in low) or ("pencil" in low and "liner" in low):
        en = "DIRECTION FOR USE: Twist up 1-2 mm of product and apply along the lash line. Close cap firmly after use."
        fr = "MODE D’EMPLOI: Faire sortir 1 à 2 mm de produit et appliquer le long de la ligne des cils. Bien refermer le capuchon après usage."
    elif ("eye" in low or "shadow" in low or "eyeshadow" in low) and "palette" in low:
        en = "DIRECTION FOR USE: Apply to eyelids with a brush or fingertip. Blend and layer shades as desired."
        fr = "MODE D’EMPLOI: Appliquer sur les paupières à l’aide d’un pinceau ou du bout des doigts. Estomper et superposer les teintes au besoin."
    elif "lip" in low and "cheek" in low:
        en = DEFAULT_DIRECTION_EN
        fr = DEFAULT_DIRECTION_FR
    elif "lipstick" in low or "lip" in low:
        en = "DIRECTION FOR USE: Apply directly to lips. Reapply as needed."
        fr = "MODE D’EMPLOI: Appliquer directement sur les lèvres. Réappliquer au besoin."
    elif "blush" in low or "palette" in low:
        en = "DIRECTION FOR USE: Apply to cheeks with a brush. Mix, match, and layer shades as desired."
        fr = "MODE D’EMPLOI: Appliquer sur les joues à l’aide d’un pinceau. Mélanger, assortir et superposer les teintes au besoin."
    else:
        en = "DIRECTION FOR USE: Apply a proper amount to the desired area. Use as directed."
        fr = "MODE D’EMPLOI: Appliquer une quantité appropriée sur la zone souhaitée. Utiliser selon le mode d’emploi."
    if source_direction and not is_count_based_item(product, source_text) and "refriger" in source_direction.lower():
        en = (
            "DIRECTION FOR USE: Apply a small amount directly to lips. Store below 25°C "
            "and refrigerate if the product softens."
        )
        fr = (
            "MODE D’EMPLOI: Appliquer une petite quantité directement sur les lèvres. "
            "Conserver à moins de 25 °C et réfrigérer si le produit ramollit."
        )
    return en, fr


def default_cautions() -> tuple[str, str]:
    return DEFAULT_CAUTION_EN, DEFAULT_CAUTION_FR


def cautions_for_product(product: str, source_text: str = "") -> tuple[str, str]:
    if is_collectible_item(product, source_text):
        return (
            "CAUTIONS: Choking hazard. Small parts. Not for children under 3 years. "
            "For decorative use only. Keep away from heat and flame.",
            "MISES EN GARDE: Risque d’étouffement. Petites pièces. Ne convient pas aux enfants de moins de 3 ans. "
            "À usage décoratif seulement. Tenir à l’écart de la chaleur et des flammes.",
        )
    if is_tool_item(product, source_text):
        return (
            "CAUTIONS: Keep clean and dry. Do not use on irritated or damaged skin. Keep out of reach of children.",
            "MISES EN GARDE: Garder propre et sec. Ne pas utiliser sur une peau irritée ou abîmée. "
            "Garder hors de la portée des enfants.",
        )
    return default_cautions()


def check_hotlist(ingredients: str) -> tuple[list[str], str]:
    if not ingredients or ingredients == "need to review":
        return [], ""
    if is_material_label(ingredients):
        return [], ""
    text = hotlist_text().lower()
    if not text:
        return [], "Could not fetch Health Canada Hotlist; review required."
    found = []
    candidates = []
    after_prefix = re.split(r"\s+/\s+", ingredients.split(":", 1)[-1], maxsplit=1)[0]
    for item in after_prefix.split(","):
        clean = re.sub(r"\([^)]*\)", "", item).strip().lower()
        if len(clean) > 4:
            candidates.append(clean)
    for ingredient in candidates:
        pattern = rf"(?<![a-z0-9]){re.escape(ingredient)}(?![a-z0-9])"
        if re.search(pattern, text):
            found.append(ingredient.title())
    note = "Restricted/prohibited candidate found on Health Canada Hotlist." if found else ""
    return found, note


@dataclass
class FillResult:
    values: dict[str, str]
    status: str
    source_url: str
    notes: str


def empty_search_cache() -> dict[str, Any]:
    return {"version": SEARCH_CACHE_VERSION, "products": {}, "families": {}, "brands": {}}


def load_search_cache() -> dict[str, Any]:
    if not SEARCH_CACHE_PATH.exists():
        return empty_search_cache()
    try:
        data = json.loads(SEARCH_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return empty_search_cache()
    if not isinstance(data, dict):
        return empty_search_cache()
    if data.get("version") != SEARCH_CACHE_VERSION:
        return empty_search_cache()
    for key in ["products", "families", "brands"]:
        if not isinstance(data.get(key), dict):
            data[key] = {}
    return data


def save_search_cache(cache: dict[str, Any]) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    SEARCH_CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def product_cache_key(barcode: str, product: str) -> str:
    barcode = normalize_barcode(barcode)
    clean = searchable_text(normalize_product_name(product))
    return f"barcode:{barcode}" if barcode else f"product:{clean}"


def cached_fill_result(barcode: str, product: str, use_cache: bool) -> FillResult | None:
    if not use_cache:
        return None
    key = product_cache_key(barcode, product)
    with CACHE_LOCK:
        cache = load_search_cache()
        entry = cache.get("products", {}).get(key)
    if not isinstance(entry, dict):
        return None
    values = repair_label_values(dict(entry.get("values") or {}))
    ingredients = values.get("ingredients/ingrédients") or values.get("ingredients/ingr茅dients")
    if ingredients_label(ingredients) == "need to review":
        return None
    source_url = str(entry.get("source_url") or "")
    if source_url_is_missing(source_url):
        return None
    return FillResult(values, status_from_values(values, source_url), source_url, "Used cached product result.")


def cached_family_context(product: str, use_cache: bool) -> dict[str, Any] | None:
    if not use_cache:
        return None
    family_key = cacheable_family_key(product)
    if not family_key:
        return None
    with CACHE_LOCK:
        cache = load_search_cache()
        entry = cache.get("families", {}).get(family_key)
    return entry if isinstance(entry, dict) else None


def update_search_cache(barcode: str, product: str, result: FillResult, use_cache: bool) -> None:
    if not use_cache or not result or source_url_is_missing(result.source_url):
        return
    ingredients = result.values.get("ingredients/ingrédients") or result.values.get("ingredients/ingr茅dients")
    if ingredients_label(ingredients) == "need to review":
        return
    product_key = product_cache_key(barcode, product)
    family_key = cacheable_family_key(product)
    brand = product_brand(product)
    with CACHE_LOCK:
        cache = load_search_cache()
        cache["products"][product_key] = {
            "values": result.values,
            "source_url": result.source_url,
            "status": result.status,
        }
        if family_key:
            cache["families"][family_key] = merge_family_context(
                cache["families"].get(family_key),
                family_context_from_result(result),
            )
        if brand:
            remembered = cache["brands"].get(brand, [])
            if not isinstance(remembered, list):
                remembered = []
            for url in source_urls_from_text(result.source_url):
                domain = domain_from_url(url)
                if domain and domain not in remembered:
                    remembered.append(domain)
            cache["brands"][brand] = remembered[:10]
        save_search_cache(cache)


def fill_from_reference(
    ref_ws: openpyxl.worksheet.worksheet.Worksheet,
    ref_headers: dict[str, int],
    barcode: str,
) -> dict[str, str] | None:
    barcode_col = find_column(ref_headers, "barcode")
    if not barcode_col:
        return None
    for row in range(2, ref_ws.max_row + 1):
        ref_barcode = str(ref_ws.cell(row, barcode_col).value or "").strip()
        if ref_barcode and ref_barcode == barcode:
            values = {}
            for header, col in ref_headers.items():
                value = ref_ws.cell(row, col).value
                if value not in (None, ""):
                    values[header] = str(value)
            return values
    return None


def normalize_barcode(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else format(value, "f").rstrip("0").rstrip(".")
    text = str(value).strip()
    if not text:
        return ""
    text = text.replace("\u00a0", "").replace(" ", "")
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    if re.fullmatch(r"\d+\.00+", text):
        return text.split(".", 1)[0]
    if re.fullmatch(r"\d+\.0{1,8}", text):
        return text.split(".", 1)[0]
    if re.search(r"[eE][+-]?\d+$", text):
        try:
            number = Decimal(text)
            if number == number.to_integral_value():
                return str(number.quantize(Decimal(1)))
        except InvalidOperation:
            pass
    return text


def sku_from_url(url: str) -> str | None:
    patterns = [
        r"\bpid\.?(\d{6,14})\b",
        r"/(?:default/)?(\d{8,14})\.html\b",
        r"/products?/[^/?#]*?(\d{8,14})(?:[/?#.-]|$)",
    ]
    for pattern in patterns:
        match = re.search(pattern, url, flags=re.I)
        if match:
            return normalize_barcode(match.group(1))
    return None


def sku_from_text(text: str) -> str | None:
    if not text:
        return None
    patterns = [
        r"(?:barcode|bar\s*code|upc|ean|jan|gtin|sku|item\s*(?:no|number|#)|product\s*(?:code|id|number)|商品番号|品番|제품\s*번호|상품\s*번호|품번)\s*[:#：]?\s*([A-Z0-9-]{5,24})",
        r"\b(45\d{11})\b",
        r"\b(49\d{11})\b",
        r"\b(88\d{11})\b",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.I):
            value = normalize_barcode(match.group(1).strip("-_. "))
            if re.fullmatch(r"[A-Z0-9-]{5,24}", value, flags=re.I):
                return value
    return None


def inferred_sku(barcode: str, source_url: str, source_text: str) -> str:
    existing = normalize_barcode(barcode)
    if existing:
        return existing
    for url in source_urls_from_text(source_url):
        sku = sku_from_url(url)
        if sku:
            return sku
    return sku_from_text(source_text) or ""


def normalize_barcode_columns(wb: openpyxl.Workbook) -> None:
    for ws in wb.worksheets:
        headers = normalized_headers(ws)
        barcode_col = find_column(headers, "barcode")
        if not barcode_col:
            continue
        ws.cell(1, barcode_col).number_format = "@"
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx, barcode_col)
            barcode = normalize_barcode(cell.value)
            if barcode:
                cell.value = barcode
            cell.number_format = "@"


def fill_from_builtin_reference(barcode: str) -> dict[str, str] | None:
    return builtin_reference_data().get(normalize_barcode(barcode))


def repair_french_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    replacements = {
        "D鈥橢MPLOI": "D’EMPLOI",
        "INGR脡DIENTS": "INGRÉDIENTS",
        "DISTRIBU脡": "DISTRIBUÉ",
        "Fabriqu茅": "Fabriqué",
        "Cor茅e": "Corée",
        "Ch猫ne": "Chêne",
        "鈥櫭ヽ": "’éc",
        "鈥檃": "’a",
        "鈥檜": "’u",
        "鈥檌": "’i",
        "鈥檕": "’o",
        "鈥檈": "’e",
        "鈥橢": "’E",
        "鈥": "’",
        "脡": "É",
        "脠": "È",
        "脢": "Ê",
        "脿": "à",
        "茅": "é",
        "猫": "è",
        "掳": "°",
    }
    repaired = value
    for bad, good in replacements.items():
        repaired = repaired.replace(bad, good)
    repaired = re.sub(r"\bMODE D[’']EMPLOI\s*:\s*", "MODE D’EMPLOI: ", repaired)
    repaired = re.sub(r"\bMISES EN GARDE\s*:\s*", "MISES EN GARDE: ", repaired)
    repaired = re.sub(
        r"\bINGREDIENTS/(?:INGR[ÉE]?.?DIENTS|INGREDIENTS IN FRENCH)\s*:\s*",
        f"{INGREDIENTS_OUTPUT_PREFIX} ",
        repaired,
    )
    repaired = re.sub(r"\bDISTRIBUTED BY / DISTRIBUÉ PAR\s*:\s*", "DISTRIBUTED BY / DISTRIBUÉ PAR: ", repaired)
    repaired = re.sub(r"\s+", " ", repaired).strip()
    return repaired


def repair_label_values(values: dict[str, str]) -> dict[str, str]:
    repaired = {key: repair_french_text(value) for key, value in values.items()}
    for key, value in list(repaired.items()):
        if key.lower().startswith("ingredients/") and value not in (None, "", "need to review"):
            if not is_material_label(value):
                repaired[key] = ingredients_label(value)
    return repaired


def reference_source_url(
    barcode: str,
    product: str,
    reference_values: dict[str, str],
    family_context: dict[str, Any] | None = None,
) -> str:
    for key in ["source url", "source_url", "reference url", "reference_url"]:
        source = str(reference_values.get(key) or "").strip()
        if source:
            return source

    known = known_product_fallback(barcode, product)
    if known.get("source_url"):
        return known["source_url"]

    if family_context and not source_url_is_missing(family_context.get("source_url", "")):
        return family_context["source_url"]

    clean = searchable_text(product or reference_values.get("product name", ""))
    for required_tokens, source in FWEE_REFERENCE_SOURCE_RULES:
        if all(token in clean for token in required_tokens):
            return source

    urls = candidate_urls(barcode, product or reference_values.get("product name", ""))
    if urls:
        return format_source_urls(urls)
    return "need to review"


def missing_required_fields(values: dict[str, str]) -> list[str]:
    missing = []
    for field in REQUIRED_LABEL_FIELDS:
        value = values.get(field)
        if value in (None, "", "need to review"):
            missing.append(field)
    return missing


def is_input_row_blank(barcode: str, product: str) -> bool:
    return not str(barcode or "").strip() and not str(product or "").strip()


def clear_generated_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    headers: dict[str, int],
    row_idx: int,
) -> None:
    generated_headers = REQUIRED_LABEL_FIELDS + [
        "source url",
    ]
    for header in generated_headers:
        col = find_column(headers, header)
        if col:
            ws.cell(row_idx, col).value = None


def candidate_urls(
    barcode: str,
    product: str,
    search_mode: str = DEFAULT_SEARCH_MODE,
    use_cache: bool = True,
) -> list[str]:
    clean_product = normalize_product_name(product)
    clean_lookup = " ".join(searchable_text(name) for name in expanded_product_names(clean_product))
    candidates: list[str] = []
    if barcode == "1129343972":
        candidates.extend(
            [
                "https://www.yesstyle.com/en/into-you-glowing-lipstick-8-colors-gl08-red-brown-3g/info.html/pid.1129343972",
                "https://www.asianbeautywholesale.com/en/into-you-glowing-lipstick-8-colors-gl08-red-brown-3g/info.html/pid.1129343972",
                "https://www.intoyoucosmetics.com/en-ca/products/into-you-glow-lipstick?variant=57958599524655",
                "https://www.intoyoucosmetics.com/en-ca/products/into-you-glow-lipstick?variant=57958599590191",
                "https://www.uniquebunny.com/products/into-you-glow-lipstick",
            ]
        )
    if barcode == "1126245093":
        candidates.extend(
            [
                "https://www.yesstyle.com/en/into-you-airy-lip-cheek-mud/info.html/pid.1126245093",
                "https://www.asianbeautywholesale.com/en/into-you-airy-lip-cheek-mud/info.html/pid.1126245093",
                "https://www.uniquebunny.com/fr/products/into-you-airy-lip-cheek-mud",
            ]
        )
    if "into" in clean_lookup and "glow" in clean_lookup and "lipstick" in clean_lookup:
        candidates.append("https://www.intoyoucosmetics.com/en-ca/products/into-you-glow-lipstick?variant=57958599524655")
        candidates.append("https://www.intoyoucosmetics.com/en-ca/products/into-you-glow-lipstick?variant=57958599590191")
        candidates.append("https://www.uniquebunny.com/products/into-you-glow-lipstick")
    if "into" in clean_lookup and "airy" in clean_lookup and "lip" in clean_lookup:
        candidates.append("https://www.intoyoucosmetics.com/en-ca/products/airy-lip-mud")
        candidates.append("https://www.uniquebunny.com/fr/products/into-you-airy-lip-cheek-mud")
    if "into" in clean_lookup and "shero" in clean_lookup and "lip" in clean_lookup and "mud" in clean_lookup:
        candidates.extend(
            [
                "https://www.intoyoucosmetics.com/en-ca/products/into-you-mini-shero-lip-mud-1g-0-03oz",
                "https://www.uniquebunny.com/fr/products/into-you-shero-super-matte-lip-cheek-mud",
                "https://www.uniquebunny.com/products/into-you-shero-super-matte-lip-cheek-mud",
                "https://www.yesstyle.com/en/into-you-shero-super-matte-lip-cheek-mud-canned-9-colors-342-dust/info.html/pid.1126025802",
                "https://www.asianbeautywholesale.com/en/into-you-shero-super-matte-lip-cheek-mud-canned-9-colors-342-dust/info.html/pid.1126025802",
                "https://www.kiseki.ca/intoyou-shero-super-matte-lip-cheek-mud-em10.html",
            ]
        )
    if "3ce" in clean_lookup and "mood" in clean_lookup and "recipe" in clean_lookup and "lip" in clean_lookup:
        candidates.extend(
            [
                "https://www.3cecosmetics.com/all-products/lips/lipstick/3ce-mood-recipe-matte-lip-color",
                "https://incidecoder.com/products/3ce-mood-recipe-lip-color",
                "https://www.hwahae.com/en/products/1795698",
            ]
        )
    if (
        barcode == "521678"
        or ("cushion" in clean_lookup and "foundation" in clean_lookup and ("&be" in clean_product.lower() or " be " in f" {clean_lookup} "))
    ):
        candidates.extend(
            [
                "https://www.yesstyle.com/en/be-cushion-foundation-light-beige-spf-24-pa-12g/info.html/pid.1124871719",
                "https://www.japaniful.com/en-ca/collections/be/products/be-cushion-foundation-12g",
                "https://www.hwahae.com/en/products/be-Cushion-Foundation-SPF24-PAPLUS-PLUS-PLUS-Beige/2165908/ingredients",
            ]
        )
    if barcode == "8809030733368" or ("abib" in clean_lookup and "mild" in clean_lookup and "acidic" in clean_lookup and ("aqua" in clean_lookup or "fit" in clean_lookup)):
        candidates.extend(
            [
                "https://en.abib.com/products/aqua-fit",
                "https://www.yesstyle.com/en/abib-mild-acidic-ph-sheet-mask-aqua-fit-10-pcs/info.html",
            ]
        )
    if "3ce" in clean_lookup and "slim" in clean_lookup and "velvet" in clean_lookup and "lip" in clean_lookup:
        candidates.extend(
            [
                "https://www.yesstyle.com/en/3ce-slim-velvet-lip-color-15-colors-true-red-new-version/info.html/pid.1068651074",
                "https://www.asianbeautywholesale.com/en/3ce-slim-velvet-lip-color-15-colors-true-red-new-version/info.html/pid.1068651074",
                "https://www.yesstyle.com/en/3ce-slim-velvet-lip-color-mood-for-blossom-edition-5-colors-hold-on/info.html/pid.1072730195",
                "https://www.hwahae.com/en/products/3CE-Slim-Velvet-Lip-Color-Cotton-Up/1827272/ingredients",
            ]
        )
    if "into" in clean_lookup and "six" in clean_lookup and "blush" in clean_lookup:
        candidates.append("https://www.yesstyle.com/en/into-you-six-color-blush-palette-six-color-blush-palette-15g/info.html/pid.1137202898")
    if "judydoll" in clean_lookup and "liquid" in clean_lookup and "blush" in clean_lookup:
        candidates.extend(
            [
                "https://judydoll.com/products/liquid-blush-serum",
                "https://www.yesstyle.com/en/judydoll-liquid-blush-serum-4-colors-01-fig-5g/info.html/pid.1136648925",
                "https://www.kiseki.ca/judydoll-liquid-blush.html",
                "https://joybeautyhub.shop/products/liquid-blush-serum",
            ]
        )
    if "millefee" in clean_lookup and "idol" in clean_lookup and "highlighter" in clean_lookup:
        candidates.append("https://millefee.com/products/idol-highlighter-palette")
        if "rose" in clean_lookup or barcode == "1137196649":
            candidates.append("https://www.yesstyle.com/en/millefee-idol-highlighter-palette-02-rose-pink/info.html/pid.1137196649")
        if "ice" in clean_lookup or barcode == "1137196647":
            candidates.append("https://www.yesstyle.com/en/millefee-idol-highlighter-palette-01-ice-blue/info.html/pid.1137196647")
    if "fwee" in clean_lookup and "glowy" in clean_lookup and "jelly" in clean_lookup:
        candidates.extend(
            [
                "https://fwee.us/products/jelly-pot",
                "https://www.ulta.com/p/lip-cheek-glowy-jelly-pot-pimprod2053366",
            ]
        )
    if "fwee" in clean_lookup and "rose" in clean_lookup and "stay" in clean_lookup and "tint" in clean_lookup:
        candidates.extend(
            [
                "https://www.yesstyle.com/en/fwee-rose-obsession/info.html/pid.1136684590",
                "https://www.hwahae.com/en/products/fwee-Rose-Obsession-Stay-fit-Lip-Tint-GW02-Spring-Rose/2179463",
            ]
        )
    if "fwee" in clean_lookup and "3d" in clean_lookup and "voluming" in clean_lookup and "gloss" in clean_lookup:
        candidates.extend(
            [
                "https://fwee.us/products/3d-voluming-gloss",
                "https://www.ulta.com/p/3d-voluming-gloss-70-pimprod2053324",
            ]
        )
    if "fwee" in clean_lookup and "pocket" in clean_lookup and "eye" in clean_lookup and "palette" in clean_lookup:
        candidates.extend(
            [
                "https://fwee.us/products/pocket-eye-palette-1",
                "https://www.ulta.com/p/pocket-eyeshadow-palette-pimprod2053338",
            ]
        )
    if "fwee" in clean_lookup and "silicone" in clean_lookup and ("jumbo" in clean_lookup or "applicator" in clean_lookup):
        candidates.append("https://fwee.us/products/jumbo-silicone-jumbo-makeup-applicator")
    if "smiski" in clean_lookup:
        if "bath" in clean_lookup:
            candidates.extend(
                [
                    "https://smiski.com/e/products/bath/",
                    "https://www.smiskifigures.com/product/smiski-bath-series/",
                    "https://www.littleobsessed.com/smiski-bath-series-blind-box.html",
                    "https://www.japanla.com/products/smiski-bath-series-blind-box",
                    "https://shumistore.com/products/smiski-bath-series-blind-box",
                ]
            )
        elif "figure" in clean_lookup or "blind" in clean_lookup:
            candidates.extend(
                [
                    "https://smiski.com/e/products/",
                    "https://www.smiskifigures.com/",
                    "https://www.littleobsessed.com/search?q=smiski",
                    "https://www.japanla.com/search?q=smiski",
                ]
            )
    if is_love_liner_cream_fit(clean_product):
        candidates.extend(
            [
                "https://www.msh-labo.com/c/make-up/eyeliner/1112",
                "https://onlineshop.japanmart.co.nz/products/love-liner-cream-fit-pencil-r-slimoval-mbr-eye-liner-0-05g",
                "https://www.hwahae.com/en/products/LOVELiner-Cream-Fit-Pencil-Liner-Ultra-Slim-Medium-Brown/2183519/ingredients",
                "https://www.yesstyle.com/en/msh-love-liner-cream-fit-pencil-medium-brown-pokemon-limited-edition/info.html/pid.1107147858",
                "https://wcosmetics.com.au/products/love-liner-cream-fit-pencil-r",
            ]
        )
    if barcode == "4570159423723" or is_love_liner_liquid_eyeliner(clean_product):
        candidates.extend(
            [
                "https://www.msh-labo.com/c/loveliner/1160",
                "https://www.samurai-drugstore.jp/default/4570159423723.html",
                "https://suzykirei.com/products/381753",
                "https://www.dodoskin.com/products/love-liner-liquid-eyeliner-r5-0-55ml-6-colors-ultra-slim-2-types",
            ]
        )

    if use_cache:
        family_entry = cached_family_context(clean_product, True)
        if family_entry:
            candidates = source_urls_from_text(family_entry.get("source_url", "")) + candidates

    scored: list[tuple[float, str]] = []
    for url in candidates:
        scored.append((fuzzy_source_score(url, url, barcode, clean_product), url))

    strong_static_candidates = any(
        exact_product_url(url, clean_product) or product_detail_bonus(url) >= 6 for url in candidates
    )
    brand_domains = brand_domains_for_product(clean_product)
    if search_mode == "Fast" and strong_static_candidates:
        search_limit = 0
    elif len(candidates) >= 4 and strong_static_candidates:
        search_limit = 0
    elif strong_static_candidates:
        search_limit = 8
    elif brand_domains:
        search_limit = 8 if search_mode == "Fast" else 18
    else:
        search_limit = 10 if search_mode == "Fast" else 30
    if search_mode == "Deep" and len(candidates) >= 4 and strong_static_candidates:
        search_limit = 8
    for query in fuzzy_queries(barcode, clean_product, search_mode)[:search_limit]:
        for result in search_web(query):
            url = result["url"]
            score = fuzzy_source_score(url, result.get("title", ""), barcode, clean_product)
            if score >= 7 or exact_product_url(url, clean_product):
                scored.append((score, url))

    deduped: list[str] = []
    seen = set()
    for _score, url in sorted(scored, key=lambda item: (-item[0], source_rank(item[1]))):
        if not url or url in seen:
            continue
        seen.add(url)
        deduped.append(url)
    return deduped[:6 if search_mode == "Fast" else 10]


def title_case_product_slug(slug: str) -> str:
    words = []
    for word in re.split(r"\s+", slug.strip()):
        if not word:
            continue
        if re.fullmatch(r"[A-Z]{2,}\d{0,3}|[A-Z]{1,3}\d{1,3}|\d+(?:\.\d+)?(?:g|ml|pcs)", word, flags=re.I):
            words.append(word.upper().replace("ML", "mL"))
        elif word.lower() in {"ph", "spf", "pa"}:
            words.append(word.upper())
        else:
            words.append(word[:1].upper() + word[1:])
    return " ".join(words)


def product_name_from_url(url: str) -> str | None:
    parsed = urlparse(url)
    segments = [unquote(segment) for segment in parsed.path.split("/") if segment]
    if not segments:
        return None
    language_segments = {"en", "en-ca", "en-us", "en-gb", "fr", "ja", "ko", "products", "product"}
    slug = ""
    if "info.html" in segments:
        idx = segments.index("info.html")
        if idx > 0:
            slug = segments[idx - 1]
    elif "products" in segments:
        idx = segments.index("products")
        if idx + 1 < len(segments):
            slug = segments[idx + 1]
    if not slug:
        for segment in reversed(segments):
            if segment.lower() not in language_segments and not segment.lower().startswith("pid."):
                slug = segment
                break
    slug = re.sub(r"\.(?:html?|php)$", "", slug, flags=re.I)
    slug = re.sub(r"\bpid\.?\d+\b", "", slug, flags=re.I)
    slug = re.sub(r"[-_+]+", " ", slug)
    slug = re.sub(r"\s+", " ", slug).strip(" /-_.")
    if len(slug) < 4 or slug.lower() in language_segments or re.fullmatch(r"\d{4,}", slug):
        return None
    return title_case_product_slug(slug)


def product_name_from_text(text: str) -> str | None:
    if not text:
        return None
    first_screen = re.sub(r"\s+", " ", text[:1400])
    split_pattern = r"\s*(?:\||｜|–|—| - |Skip to Content|Skip to content|Welcome to)\s*"
    bad_terms = re.compile(
        r"privacy|shipping|policy|login|ranking|category|brand|cart|account|menu|official site|"
        r"free shipping|new arrivals|best selling",
        flags=re.I,
    )
    product_terms = re.compile(
        r"lip|liner|eyeliner|mascara|mask|foundation|cushion|blush|palette|gloss|tint|pencil|"
        r"cream|serum|sunscreen|powder|concealer|highlighter|shadow",
        flags=re.I,
    )
    candidates = []
    for chunk in re.split(split_pattern, first_screen):
        clean = html.unescape(chunk)
        clean = re.sub(r"【([^】]+)】", r"\1 ", clean)
        clean = re.sub(r"\([^)]*(?:official|authentic|shipping|sale)[^)]*\)", " ", clean, flags=re.I)
        clean = re.sub(r"[^A-Za-z0-9#&+()./'\-\s]", " ", clean)
        clean = re.sub(r"^.*\b(?:drugstore|store|shop)\b\s+", "", clean, flags=re.I)
        clean = re.sub(r"^\s*msh\b", "MSH", clean, flags=re.I)
        clean = re.sub(r"\s+", " ", clean).strip(" .:-")
        words = re.findall(r"[A-Za-z0-9#&+()./'-]+", clean)
        if len(words) < 4 or len(clean) > 120:
            continue
        if bad_terms.search(clean) or not product_terms.search(clean):
            continue
        score = len(words) + (4 if re.search(r"\b(?:black|brown|pink|beige|red|rose|coral)\b", clean, flags=re.I) else 0)
        candidates.append((score, clean))
    if not candidates:
        return None
    return max(candidates, key=lambda item: item[0])[1]


def inferred_product_name(product: str, source_url: str, source_text: str = "") -> str | None:
    if normalize_product_name(product):
        return normalize_product_name(product)
    name = product_name_from_text(source_text)
    if name:
        return name
    for url in source_urls_from_text(source_url):
        name = product_name_from_url(url)
        if name:
            return name
    return None


def known_product_fallback(barcode: str, product: str) -> dict[str, str]:
    known = KNOWN_ONLINE_PRODUCTS.get(barcode)
    if known:
        return known
    clean = searchable_text(expand_product_abbreviations(product))
    if "into" in clean and "glow" in clean and "lipstick" in clean:
        return KNOWN_ONLINE_PRODUCTS["1129343972"]
    if "into" in clean and "airy" in clean and "lip" in clean and ("cheek" in clean or "mud" in clean):
        return KNOWN_ONLINE_PRODUCTS["1126245093"]
    if "into" in clean and "six" in clean and "blush" in clean and "palette" in clean:
        return KNOWN_ONLINE_PRODUCTS["1137202898"]
    if "3ce" in clean and "mood" in clean and "recipe" in clean and "lip" in clean and "color" in clean:
        return {
            "source_url": (
                "https://www.3cecosmetics.com/all-products/lips/lipstick/3ce-mood-recipe-matte-lip-color\n"
                "https://www.hwahae.com/en/products/1795698"
            ),
            "net weight": "Net. 3.5 g",
            "coo": "Made In Korea / Fabriqué En Corée",
        }
    if barcode == "8809030733368" or ("abib" in clean and "mild" in clean and "acidic" in clean and ("aqua" in clean or "fit" in clean)):
        return {
            "source_url": "https://en.abib.com/products/aqua-fit",
            "coo": "Made In Korea / Fabriqué En Corée",
        }
    if barcode == "4570159423723" or is_love_liner_liquid_eyeliner(product):
        return dict(LOVE_LINER_LIQUID_EYELINER_R5)
    if is_love_liner_cream_fit(product):
        known = dict(LOVE_LINER_CREAM_FIT_PENCIL_R)
        if love_liner_variant(product) != "Ultra Slim":
            known["net weight"] = "Net. 0.1 g"
        return known
    if "judydoll" in clean and "liquid" in clean and "blush" in clean:
        shade = shade_key(product)
        ingredients = JUDYDOLL_LIQUID_BLUSH_SERUM["shades"].get(shade)
        if ingredients:
            return {
                "source_url": JUDYDOLL_LIQUID_BLUSH_SERUM["source_url"],
                "net weight": JUDYDOLL_LIQUID_BLUSH_SERUM["net weight"],
                "source_direction": JUDYDOLL_LIQUID_BLUSH_SERUM["source_direction"],
                "coo": JUDYDOLL_LIQUID_BLUSH_SERUM["coo"],
                "ingredients": ingredients,
            }
    if "millefee" in clean and "idol" in clean and "highlighter" in clean:
        shade = MILLEFEE_IDOL_HIGHLIGHTER_PALETTE["barcodes"].get(str(barcode).strip())
        shade = shade or shade_key(product)
        shade_data = MILLEFEE_IDOL_HIGHLIGHTER_PALETTE["shades"].get(shade)
        if shade_data:
            return {
                "source_url": shade_data["source_url"],
                "net weight": MILLEFEE_IDOL_HIGHLIGHTER_PALETTE["net weight"],
                "source_direction": MILLEFEE_IDOL_HIGHLIGHTER_PALETTE["source_direction"],
                "coo": MILLEFEE_IDOL_HIGHLIGHTER_PALETTE["coo"],
                "ingredients": shade_data["ingredients"],
            }
    return {}


def enough_product_data(texts: list[tuple[str, str]], product: str, known: dict[str, str]) -> bool:
    if len(texts) >= 4:
        return True
    combined = " ".join(text for _url, text in texts)
    has_net = bool(
        net_weight_from_text(combined)
        or known.get("net weight")
        or net_weight_from_name(product)
        or pcs_count_from_text(product, combined)
    )
    has_ingredients = bool(
        ingredients_from_text(combined, product)
        or (ingredients_label_from_known(known) != "need to review")
        or material_label_from_text(product, combined)
    )
    if has_ingredients and texts:
        return True
    if len(texts) < 2:
        return False
    return has_net and has_ingredients


def verified_known_product(known: dict[str, str]) -> bool:
    return bool(
        known.get("source_url")
        and known.get("net weight")
        and (known.get("ingredients") or known.get("ingredients_fr"))
    )


def source_url_is_missing(source_url: str) -> bool:
    value = str(source_url or "").strip().lower()
    return value in {"", "need to review", "built-in reference data"}


def source_urls_from_text(source_url: str) -> list[str]:
    urls = []
    for line in str(source_url or "").splitlines():
        url = line.strip()
        if url.startswith("http"):
            urls.append(url)
    return urls


def format_source_urls(urls: list[str], limit: int = MAX_SOURCE_URLS) -> str:
    deduped: list[str] = []
    seen = set()
    for url in urls:
        clean = str(url or "").strip()
        if not clean or not clean.startswith("http") or clean in seen:
            continue
        seen.add(clean)
        deduped.append(clean)
        if len(deduped) >= limit:
            break
    return "\n".join(deduped)


def fetch_source_texts(
    urls: list[str],
    product: str,
    known: dict[str, str],
    search_mode: str = DEFAULT_SEARCH_MODE,
) -> list[tuple[str, str]]:
    texts: list[tuple[str, str]] = []
    if not urls:
        return texts
    first_url = urls[0]
    first_text = fetch_text(first_url)
    if len(first_text) > 200:
        texts.append((first_url, first_text))
        if enough_product_data(texts, product, known):
            return texts

    max_urls = 3 if search_mode == "Fast" else 6
    target_urls = [url for url in urls[1: min(max_urls, len(urls))] if url != first_url]
    if not target_urls:
        return texts
    if search_mode == "Fast":
        for url in target_urls:
            text = fetch_text(url)
            if len(text) > 200:
                texts.append((url, text))
                if enough_product_data(texts, product, known):
                    break
        return texts
    max_workers = min(4, len(target_urls))
    executor = ThreadPoolExecutor(max_workers=max_workers)
    try:
        future_to_url = {executor.submit(fetch_text, url): url for url in target_urls}
        for future in as_completed(future_to_url):
            url = future_to_url[future]
            try:
                text = future.result()
            except Exception:
                continue
            if len(text) > 200:
                texts.append((url, text))
                if enough_product_data(texts, product, known):
                    for pending in future_to_url:
                        if not pending.done():
                            pending.cancel()
                    break
    finally:
        executor.shutdown(wait=False, cancel_futures=True)
    return sorted(texts, key=lambda item: urls.index(item[0]) if item[0] in urls else len(urls))


def status_from_values(values: dict[str, str], source_url: str) -> str:
    missing = [
        field
        for field in REQUIRED_LABEL_FIELDS
        if values.get(field) in (None, "", "need to review")
    ]
    restricted = bool(values.get("restricted ingredients"))
    if source_url_is_missing(source_url):
        return "Missing source"
    return "Need to review" if missing or restricted else "Completed"


def apply_family_context(result: FillResult, family_context: dict[str, Any] | None) -> FillResult:
    if not family_context:
        return result
    shared_values = family_context.get("values", {})
    for field, value in shared_values.items():
        if value and value != "need to review" and result.values.get(field) in (None, "", "need to review"):
            result.values[field] = value
    if source_url_is_missing(result.source_url) and not source_url_is_missing(family_context.get("source_url", "")):
        result.source_url = family_context["source_url"]
    result.status = status_from_values(result.values, result.source_url)
    return result


def family_context_from_result(result: FillResult) -> dict[str, Any]:
    values = {
        field: result.values.get(field)
        for field in SHARED_FAMILY_FIELDS
        if result.values.get(field) not in (None, "", "need to review")
    }
    source_url = "" if source_url_is_missing(result.source_url) else result.source_url
    return {"values": values, "source_url": source_url}


def merge_family_context(existing: dict[str, Any] | None, new: dict[str, Any]) -> dict[str, Any]:
    if not existing:
        return new
    merged = {"values": dict(existing.get("values", {})), "source_url": existing.get("source_url", "")}
    for field, value in new.get("values", {}).items():
        if value and value != "need to review":
            merged["values"][field] = value
    if new.get("source_url") and not source_url_is_missing(new.get("source_url", "")):
        existing_urls = source_urls_from_text(merged.get("source_url", ""))
        for url in source_urls_from_text(new["source_url"]):
            if url not in existing_urls:
                existing_urls.append(url)
        merged["source_url"] = format_source_urls(existing_urls)
    return merged


def process_row(
    row: dict[str, Any],
    reference_values: dict[str, str] | None,
    use_defaults: bool,
    family_context: dict[str, Any] | None = None,
    search_mode: str = DEFAULT_SEARCH_MODE,
    use_cache: bool = True,
) -> FillResult:
    product = str(row.get("product name") or "")
    barcode = normalize_barcode(row.get("barcode"))
    values: dict[str, str] = {}
    source_url = ""
    notes: list[str] = []

    if reference_values:
        values.update(reference_values)
        values = repair_label_values(values)
        source_product = product or values.get("product name", "")
        if source_product and french_name_needs_translation(values.get("product name french", "")):
            values["product name french"] = product_name_fr(source_product)
        notes.append("Matched reference data by barcode.")
        source_url = reference_source_url(barcode, product, reference_values, family_context)
        missing = missing_required_fields(values)
        for field in missing:
            values[field] = "need to review"
        if missing:
            notes.append("Missing required fields: " + ", ".join(missing) + ".")
        if source_url_is_missing(source_url):
            notes.append("Reference data has no source URL; manual source review required.")
        else:
            notes.append("Added source URL for matched reference data.")
        return apply_family_context(FillResult(
            values,
            status_from_values(values, source_url),
            source_url,
            " ".join(notes),
        ), family_context)

    cached = cached_fill_result(barcode, product, use_cache)
    if cached and (barcode or cached.values.get("barcode")):
        return cached

    known = known_product_fallback(barcode, product)
    texts: list[tuple[str, str]] = []
    source_candidates: list[str] = source_urls_from_text(known.get("source_url", ""))
    if verified_known_product(known):
        source_url = known["source_url"]
        notes.append("Used verified product data for this shade.")
        if not normalize_product_name(product):
            for url in source_urls_from_text(source_url):
                text = fetch_text(url)
                if len(text) > 200:
                    texts.append((url, text))
                    if product_name_from_text(text):
                        break
    else:
        if not family_context:
            family_context = cached_family_context(product, use_cache)
        urls = source_urls_from_text(family_context.get("source_url", "")) if family_context else []
        if urls:
            notes.append("Reused source URLs from matching product family.")
        else:
            urls = candidate_urls(barcode, product, search_mode, use_cache)
        source_candidates = urls + source_candidates
        texts = fetch_source_texts(urls, product, known, search_mode)
        if texts:
            checked_urls = [url for url, _text in texts]
            ingredient_urls = [url for url, text in texts if ingredients_from_text(text, product)]
            source_url = format_source_urls(
                ingredient_urls + checked_urls + source_candidates,
            )
            notes.append("Sources checked: " + ", ".join(domain_from_url(url) for url, _text in texts[:4]))
            remaining = len(source_urls_from_text(source_url)) - len(checked_urls)
            if remaining > 0:
                notes.append("Additional candidate source URLs included for manual review.")
        elif source_candidates:
            source_url = format_source_urls(source_candidates)
            notes.append("Candidate source URLs found; extraction incomplete.")
        else:
            notes.append("No reliable source found.")

    if known and not source_url:
        source_url = known.get("source_url", "")
        notes.append("Used approved source URLs for this SKU.")
    source_url = format_source_urls(source_urls_from_text(source_url))

    combined_text = " ".join(text for _url, text in texts)
    resolved_barcode = inferred_sku(barcode, source_url, combined_text)
    resolved_product = normalize_product_name(product)
    if not resolved_product:
        for _url, text in texts:
            resolved_product = product_name_from_text(text) or ""
            if resolved_product:
                break
    resolved_product = resolved_product or inferred_product_name(product, source_url, combined_text) or product
    source_direction = how_to_use_from_text(combined_text) or known.get("source_direction")
    direction_en, direction_fr = direction_for_product(resolved_product, source_direction, combined_text)
    caution_en, caution_fr = cautions_for_product(resolved_product, combined_text)
    count_net_weight = pcs_count_from_text(resolved_product, combined_text)

    values["barcode"] = resolved_barcode
    values["product name"] = resolved_product
    values["product name french"] = product_name_fr(resolved_product)
    net_weight = (
        compatible_net_weight(resolved_product, known.get("net weight"))
        or compatible_net_weight(resolved_product, net_weight_from_name(resolved_product))
        or compatible_net_weight(resolved_product, net_weight_from_text(combined_text))
        or compatible_net_weight(resolved_product, count_net_weight)
    )
    values["net weight"] = net_weight or "need to review"
    values["direction for use"] = direction_en
    values["mode d’emploi"] = direction_fr
    values["cautions"] = caution_en
    values["mises en garde:"] = caution_fr
    values["manufacturer"] = manufacturer_from_text(combined_text) or known.get("manufacturer") or "need to review"
    values["ingredients/ingrédients"] = (
        ingredients_from_text(combined_text, resolved_product)
        or ingredients_label_from_known(known)
        or material_label_from_text(resolved_product, combined_text)
        or "need to review"
    )
    values["coo"] = coo_from_text(combined_text + " " + resolved_product) or known.get("coo") or "need to review"
    values["distributed by / distribué par:"] = DISTRIBUTOR

    for field in REQUIRED_LABEL_FIELDS:
        if field not in values:
            values[field] = "need to review"
    values = repair_label_values(values)

    restricted, hotlist_note = check_hotlist(values.get("ingredients/ingrédients", ""))
    if restricted:
        values["restricted ingredients"] = ", ".join(restricted)
        notes.append(hotlist_note)
    elif hotlist_note:
        notes.append(hotlist_note)

    status = status_from_values(values, source_url)
    result = apply_family_context(FillResult(values, status, source_url, " ".join(notes)), family_context)
    update_search_cache(barcode, product, result, use_cache)
    return result


def dataframe_from_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, rows: int = 20) -> pd.DataFrame:
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return pd.DataFrame()
    headers = [str(h) if h is not None else f"Unnamed {i+1}" for i, h in enumerate(data[0])]
    return pd.DataFrame(data[1 : rows + 1], columns=headers)


def full_dataframe_from_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> pd.DataFrame:
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return pd.DataFrame()
    headers = [str(h) if h is not None else f"Unnamed {i+1}" for i, h in enumerate(data[0])]
    return pd.DataFrame(data[1:], columns=headers).fillna("")


def need_review_cell_css(value: Any) -> str:
    return "background-color: #fff2cc" if is_need_review_value(value) else ""


def highlighted_review_dataframe(df: pd.DataFrame) -> Any:
    styler = df.style
    if hasattr(styler, "map"):
        return styler.map(need_review_cell_css)
    return styler.applymap(need_review_cell_css)


def apply_dataframe_to_sheet(
    wb_path: Path,
    sheet_name: str,
    df: pd.DataFrame,
    output_name: str,
) -> Path:
    wb = openpyxl.load_workbook(wb_path)
    ws = wb[sheet_name]
    for col_idx, header in enumerate(df.columns, start=1):
        ws.cell(1, col_idx).value = header if not str(header).startswith("Unnamed ") else None
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            header = str(df.columns[col_idx - 1]).strip().lower()
            cell = ws.cell(row_idx, col_idx)
            if "barcode" in header:
                barcode = normalize_barcode(value)
                cell.value = barcode or None
                cell.number_format = "@"
            else:
                cell.value = None if value == "" else value
            apply_need_review_highlight(cell, clear_existing=True)
    return export_workbook(wb, output_name)


def workbook_bytes(uploaded_file: Any) -> bytes:
    return uploaded_file.getvalue()


def save_upload(uploaded_file: Any) -> Path:
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=DATA_DIR)
    tmp.write(workbook_bytes(uploaded_file))
    tmp.close()
    return Path(tmp.name)


def export_workbook(wb: openpyxl.Workbook, name: str) -> Path:
    EXPORT_DIR.mkdir(exist_ok=True)
    safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", name).strip("_") or "filled_labels.xlsx"
    output = EXPORT_DIR / safe
    for ws in wb.worksheets:
        remove_manufacturer_columns(ws)
    normalize_barcode_columns(wb)
    highlight_need_review_cells(wb)
    wb.save(output)
    return output


def process_workbook(
    path: Path,
    fill_sheet: str,
    use_defaults: bool,
    search_mode: str = DEFAULT_SEARCH_MODE,
    use_cache: bool = True,
    max_workers: int = 4,
    limit: int | None = None,
) -> tuple[openpyxl.Workbook, pd.DataFrame]:
    wb = openpyxl.load_workbook(path)
    fill_ws = wb[fill_sheet]
    fill_headers = add_audit_columns(fill_ws)

    barcode_col = find_column(fill_headers, "barcode")
    if not barcode_col:
        raise ValueError("No barcode column found.")

    product_col = find_column(fill_headers, "product name")
    records = []
    task_groups: dict[str, list[dict[str, Any]]] = {}
    max_row = fill_ws.max_row if limit is None else min(fill_ws.max_row, limit + 1)
    for row_idx in range(2, max_row + 1):
        barcode = normalize_barcode(fill_ws.cell(row_idx, barcode_col).value)
        product = str(fill_ws.cell(row_idx, product_col).value or "") if product_col else ""
        if is_input_row_blank(barcode, product):
            clear_generated_row(fill_ws, fill_headers, row_idx)
            continue
        fill_ws.cell(row_idx, barcode_col).value = barcode
        fill_ws.cell(row_idx, barcode_col).number_format = "@"
        family_key = cacheable_family_key(product)
        group_key = family_key or f"row:{row_idx}"
        task_groups.setdefault(group_key, []).append(
            {
                "row_idx": row_idx,
                "barcode": barcode,
                "product": product,
                "family_key": family_key,
                "reference_values": fill_from_builtin_reference(barcode),
            }
        )

    def process_group(group: list[dict[str, Any]]) -> list[dict[str, Any]]:
        group_results = []
        family_context = cached_family_context(group[0]["product"], use_cache) if group[0]["family_key"] else None
        for item in group:
            result = process_row(
                {"barcode": item["barcode"], "product name": item["product"]},
                item["reference_values"],
                use_defaults,
                family_context,
                search_mode,
                use_cache,
            )
            if item["family_key"]:
                family_context = merge_family_context(family_context, family_context_from_result(result))
            group_results.append({**item, "result": result})
        return group_results

    processed_items: list[dict[str, Any]] = []
    groups = list(task_groups.values())
    if max_workers <= 1 or len(groups) <= 1:
        for group in groups:
            processed_items.extend(process_group(group))
    else:
        with ThreadPoolExecutor(max_workers=min(max_workers, len(groups))) as executor:
            futures = [executor.submit(process_group, group) for group in groups]
            for future in as_completed(futures):
                processed_items.extend(future.result())

    for item in sorted(processed_items, key=lambda value: value["row_idx"]):
        row_idx = item["row_idx"]
        barcode = item["barcode"]
        product = item["product"]
        result = item["result"]
        for header, value in result.values.items():
            col = find_column(fill_headers, header)
            if col and value:
                fill_ws.cell(row_idx, col).value = value
                if row_idx > 2:
                    copy_cell_style(fill_ws.cell(row_idx - 1, col), fill_ws.cell(row_idx, col))

        fill_ws.cell(row_idx, fill_headers["source url"]).value = result.source_url
        records.append(
            {
                "row": row_idx,
                "barcode": barcode,
                "product": product,
                "status": result.status,
                "source_url": result.source_url,
                "notes": result.notes,
            }
        )

    return wb, pd.DataFrame(records)


def direct_input_rows(df: pd.DataFrame) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for raw in df.to_dict("records"):
        barcode = normalize_barcode(raw.get("barcode") or raw.get("Barcode"))
        product = normalize_product_name(raw.get("product name") or raw.get("Product Name") or "")
        if is_input_row_blank(barcode, product):
            continue
        rows.append({"barcode": barcode, "product name": product})
    return rows


def process_direct_rows(
    rows: list[dict[str, str]],
    use_defaults: bool,
    search_mode: str = DEFAULT_SEARCH_MODE,
    use_cache: bool = True,
    max_workers: int = 4,
) -> pd.DataFrame:
    task_groups: dict[str, list[dict[str, Any]]] = {}
    for idx, row in enumerate(rows, start=1):
        barcode = normalize_barcode(row.get("barcode"))
        product = normalize_product_name(row.get("product name") or "")
        family_key = cacheable_family_key(product)
        group_key = family_key or f"row:{idx}"
        task_groups.setdefault(group_key, []).append(
            {
                "idx": idx,
                "barcode": barcode,
                "product": product,
                "family_key": family_key,
                "reference_values": fill_from_builtin_reference(barcode),
            }
        )

    def process_group(group: list[dict[str, Any]]) -> list[dict[str, Any]]:
        group_results = []
        family_context = cached_family_context(group[0]["product"], use_cache) if group[0]["family_key"] else None
        for item in group:
            result = process_row(
                {"barcode": item["barcode"], "product name": item["product"]},
                item["reference_values"],
                use_defaults,
                family_context,
                search_mode,
                use_cache,
            )
            if item["family_key"]:
                family_context = merge_family_context(family_context, family_context_from_result(result))
            group_results.append({**item, "result": result})
        return group_results

    processed_items: list[dict[str, Any]] = []
    groups = list(task_groups.values())
    if max_workers <= 1 or len(groups) <= 1:
        for group in groups:
            processed_items.extend(process_group(group))
    else:
        with ThreadPoolExecutor(max_workers=min(max_workers, len(groups))) as executor:
            futures = [executor.submit(process_group, group) for group in groups]
            for future in as_completed(futures):
                processed_items.extend(future.result())

    records = []
    for item in sorted(processed_items, key=lambda value: value["idx"]):
        result = item["result"]
        resolved_product = result.values.get("product name") or item["product"]
        record = {
            "barcode": result.values.get("barcode") or item["barcode"],
            "product name": resolved_product,
        }
        for field in REQUIRED_LABEL_FIELDS:
            record[field] = result.values.get(field, "need to review")
        if result.values.get("restricted ingredients"):
            record["restricted ingredients"] = result.values["restricted ingredients"]
        record["source url"] = result.source_url
        records.append(record)
    return pd.DataFrame(records)


def dataframe_to_workbook(df: pd.DataFrame, sheet_name: str = "Direct Search") -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    for col_idx, header in enumerate(df.columns, start=1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="D9EAF7")
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            header = str(df.columns[col_idx - 1]).strip().lower()
            cell = ws.cell(row_idx, col_idx)
            if "barcode" in header:
                cell.value = normalize_barcode(value)
                cell.number_format = "@"
            else:
                cell.value = None if value == "" else value
            apply_need_review_highlight(cell, clear_existing=True)
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    for col_idx, header in enumerate(df.columns, start=1):
        header_text = str(header).lower()
        if header_text == "barcode":
            width = 16
        elif header_text in {"product name", "product name french", "net weight"}:
            width = 28
        elif "source" in header_text or "ingredients" in header_text or "direction" in header_text:
            width = 54
        else:
            width = 34
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    return wb


def export_direct_dataframe(df: pd.DataFrame, output_name: str = "direct_label_info.xlsx") -> Path:
    wb = dataframe_to_workbook(df)
    return export_workbook(wb, output_name)


def template_workbook_bytes() -> bytes:
    columns = ["barcode", "product name", *REQUIRED_LABEL_FIELDS, "source url"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Label Info Template"
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    for row_idx in range(2, 22):
        ws.cell(row_idx, 1).number_format = "@"
    widths = {
        "barcode": 18,
        "product name": 34,
        "product name french": 34,
        "net weight": 18,
        "source url": 54,
    }
    for col_idx, header in enumerate(columns, start=1):
        width = widths.get(header, 46 if any(key in header for key in ["direction", "ingredients", "cautions", "garde"]) else 30)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=DATA_DIR) as tmp:
        tmp_path = Path(tmp.name)
    try:
        wb.save(tmp_path)
        return tmp_path.read_bytes()
    finally:
        try:
            tmp_path.unlink()
        except OSError:
            pass


def login_screen() -> None:
    st.title(SITE_NAME)
    st.caption("Private workbook processing for bilingual Nakama labels.")
    with st.form("login"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Sign in")
    if submitted:
        user = authenticate(username, password)
        if user:
            st.session_state.user = user
            st.rerun()
        st.error("Invalid username or password.")


def excel_workbook_section() -> None:
    st.download_button(
        "Download Excel template",
        data=template_workbook_bytes(),
        file_name="label_info_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    uploaded = st.file_uploader("Excel workbook", type=["xlsx"])
    if not uploaded:
        return

    path = save_upload(uploaded)
    wb = openpyxl.load_workbook(path)
    names = wb.sheetnames
    fill_sheet = st.selectbox(
        "Sheet to fill",
        names,
        index=names.index("Sheet2") if "Sheet2" in names else 0,
    )

    with st.expander("Preview sheet to fill", expanded=True):
        st.dataframe(dataframe_from_sheet(wb[fill_sheet]), use_container_width=True)

    use_defaults = st.checkbox(
        "Use approved default lip/cheek direction and general cautions when source data is missing",
        value=True,
        key="workbook_use_defaults",
    )
    search_mode = st.radio(
        "Search mode",
        SEARCH_MODES,
        index=0,
        horizontal=True,
        key="workbook_search_mode",
        help="Fast checks known official/trusted sources first. Deep searches more broadly and is slower.",
    )
    use_cache = st.checkbox(
        "Use saved search cache",
        value=True,
        key="workbook_use_cache",
    )
    parallel_rows = st.slider(
        "Parallel product groups",
        min_value=1,
        max_value=6,
        value=4,
        key="workbook_parallel_rows",
        help="Higher values can be faster, but too high may trigger website timeouts.",
    )
    limit = st.number_input("Rows to process now (0 = all)", min_value=0, value=0, step=1)

    if st.button("Process workbook", type="primary"):
        with st.status("Processing product groups...", expanded=True) as status:
            processed_wb, report = process_workbook(
                path,
                fill_sheet,
                use_defaults,
                search_mode,
                use_cache,
                int(parallel_rows),
                None if limit == 0 else int(limit),
            )
            output_path = export_workbook(processed_wb, f"filled_{uploaded.name}")
            status.update(label="Processing complete", state="complete")
        st.session_state.report = report
        st.session_state.output_path = str(output_path)
        st.session_state.fill_sheet = fill_sheet
        st.session_state.uploaded_name = uploaded.name

    if "report" in st.session_state:
        st.subheader("Row status")
        st.dataframe(
            highlighted_review_dataframe(st.session_state.report),
            use_container_width=True,
            hide_index=True,
        )
        output_path = Path(st.session_state.output_path)

        st.subheader("Manual edit before export")
        processed_wb = openpyxl.load_workbook(output_path)
        edit_df = full_dataframe_from_sheet(processed_wb[st.session_state.fill_sheet])
        edited_df = st.data_editor(
            edit_df,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            key="processed_sheet_editor",
        )
        if st.button("Save edited workbook"):
            edited_path = apply_dataframe_to_sheet(
                output_path,
                st.session_state.fill_sheet,
                edited_df,
                f"edited_{st.session_state.uploaded_name}",
            )
            st.session_state.output_path = str(edited_path)
            output_path = edited_path
            st.success("Edited workbook saved.")

        st.download_button(
            "Download completed Excel",
            data=output_path.read_bytes(),
            file_name=output_path.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def direct_search_section() -> None:
    default_input = pd.DataFrame(
        [
            {"barcode": "", "product name": ""},
            {"barcode": "", "product name": ""},
        ]
    )
    input_df = st.data_editor(
        default_input,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        key="direct_input_editor",
        column_config={
            "barcode": st.column_config.TextColumn("Barcode / SKU", width="medium"),
            "product name": st.column_config.TextColumn("Product Name", width="large"),
        },
    )
    use_defaults = st.checkbox(
        "Use approved default lip/cheek direction and general cautions when source data is missing",
        value=True,
        key="direct_use_defaults",
    )
    search_mode = st.radio(
        "Search mode",
        SEARCH_MODES,
        index=0,
        horizontal=True,
        key="direct_search_mode",
        help="Fast checks known official/trusted sources first. Deep searches more broadly and is slower.",
    )
    use_cache = st.checkbox(
        "Use saved search cache",
        value=True,
        key="direct_use_cache",
    )
    parallel_rows = st.slider(
        "Parallel product groups",
        min_value=1,
        max_value=6,
        value=4,
        key="direct_parallel_rows",
    )

    if st.button("Search products", type="primary", key="direct_search_button"):
        rows = direct_input_rows(input_df)
        if not rows:
            st.warning("Enter at least one barcode/SKU or product name.")
        else:
            with st.status("Searching product information...", expanded=True) as status:
                result_df = process_direct_rows(rows, use_defaults, search_mode, use_cache, int(parallel_rows))
                output_path = export_direct_dataframe(result_df)
                status.update(label="Search complete", state="complete")
            st.session_state.direct_results = result_df
            st.session_state.direct_output_path = str(output_path)

    if "direct_results" in st.session_state:
        st.subheader("Search results")
        st.dataframe(
            highlighted_review_dataframe(st.session_state.direct_results),
            use_container_width=True,
            hide_index=True,
        )

        st.subheader("Manual edit before export")
        edited_df = st.data_editor(
            st.session_state.direct_results,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            key="direct_results_editor",
        )
        if st.button("Save edited direct Excel", key="save_direct_results"):
            output_path = export_direct_dataframe(edited_df, "edited_direct_label_info.xlsx")
            st.session_state.direct_results = edited_df
            st.session_state.direct_output_path = str(output_path)
            st.success("Edited direct Excel saved.")

        output_path = Path(st.session_state.direct_output_path)
        st.download_button(
            "Download direct Excel",
            data=output_path.read_bytes(),
            file_name=output_path.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def main_app() -> None:
    user = st.session_state.user
    st.sidebar.write(f"Signed in as **{user['username']}**")
    if st.sidebar.button("Sign out"):
        st.session_state.pop("user", None)
        st.rerun()

    page = st.sidebar.radio(
        "Navigation",
        ["Fill labels", "Manage users"] if user["role"] == "admin" else ["Fill labels"],
    )
    if page == "Manage users":
        manage_users()
        return

    st.title(SITE_NAME)
    st.caption("Process full workbooks or search one-off missing products.")
    input_mode = st.radio(
        "Input method",
        ["Upload Excel workbook", "Direct product search"],
        horizontal=True,
    )
    if input_mode == "Direct product search":
        direct_search_section()
    else:
        excel_workbook_section()


def main() -> None:
    st.set_page_config(page_title=SITE_NAME, layout="wide")
    ensure_dirs()
    db().close()
    if "user" not in st.session_state:
        login_screen()
    else:
        main_app()


if __name__ == "__main__":
    main()
